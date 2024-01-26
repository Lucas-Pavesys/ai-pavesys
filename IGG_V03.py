# ------------------------ IMPORTAÇÕES ------------------------
import os
import pandas as pd
import numpy as np
import ler_arquivo_V01 as leitor
from tqdm import tqdm
import openpyxl as op
import shutil
from copy import copy
from LVD_Pavesys_V03 import AreaTrincada
import math
from random import randint
import warnings
warnings.filterwarnings("ignore")


# ------------------------ FUNÇÕES GERAIS ------------------------
def ExtrairDF(path, pista, atr): # Extrai os DataFrames das avaliações e atr
    # Extrai DataFrames
    file_loc = os.path.join(path, pista)
    file_loc_atr = os.path.join(path, atr)
    df_original = pd.read_excel(file_loc, header = 7)
    df_original_atr = pd.read_excel(file_loc_atr, header = 19, usecols = "B:K")
    # Ajusta DataFrames
    df_original_atr.replace(to_replace=["-"], value=[None], inplace=True)
    df_original.rename(columns={'FI': 'Fi'}, inplace=True)
    return file_loc, df_original, df_original_atr


def DadosCabecalho(path, pista, atr_name, df_obs):
    # Extrai dados dos cabeçalhos dos dataframes da avaliação e do ATR
    # Dados do IGG
    df_original = pd.read_excel(os.path.join(path, pista), usecols = "A:B")
    for coluna in df_original.columns:
        if coluna != "Rodovia":
            rodovia = str(coluna)
    nome_trecho = pista.split("_")[1]
    tipo_pista = str(df_original.iat[0, 1])
    faixa = str(df_original.iat[1, 1])
    data_bruta = df_original.iat[4, 1]
    try:
        data = ("{}/{}/{}".format(data_bruta.day, data_bruta.month, data_bruta.year))
    except:
        data = data_bruta
    km_ini = float(df_original.iat[2, 1])
    km_fim = float(df_original.iat[3, 1])
    kms_trecho = "KM"+str(f'{km_ini:.3f}').replace(".","+")+" ao KM"+str(f'{km_fim:.3f}').replace(".","+")
    sentido_pista = 'Crescente' if km_ini - km_fim < 0 else 'Decrescente'
    lado_pista = 'D' if km_ini - km_fim < 0 else 'E'
    # via = df_original.iat[5, 1]
    via = ""
    # Falta criar forma de usar esse abaixo - Apenas usados em dispositivos
    km_referencia = ""
    nome_ramo = ""
    try:
        if "+" in df_obs['Observação'][0]:
            for temp in df_obs['Observação'][0].split(" | "):
                if "+" in temp:
                    km_referencia = (temp.split("/")[1]).split("-")[0]
                    nome_ramo = ((temp.split("/")[1]).split("-")[1]).split("(")[0]
                    faixa = (temp.split("Faixa")[1]).split("(")[0]
    except:
        pass
    # Dados do ATR
    operador = ""
    path_atr = os.path.join(path, atr_name)
    if os.path.isfile(path_atr):
        df_original_atr = pd.read_excel(path_atr, header = 5, usecols = "B:K")
        operador = str(df_original_atr.iat[4, 3])
    return {"Rodovia":rodovia, "STH":nome_trecho, "Tipo pista":tipo_pista, "Faixa":faixa, 
            "Sentido":sentido_pista, "Data":data, "Operador":operador, "km ini":km_ini, "km fim":km_fim,
            "Trecho":kms_trecho, "Lado": lado_pista, "Secao Terra.": "MISTA", "Lote": "", "Tipo Revest.":"CA",
            "km referencia":km_referencia, "Nome ramo":nome_ramo, "Via":via}


def ConcatenaDf(coluna, *args): # RECEBE APENAS DATAFRAME
    # Função auxiliar da: PadraoLVC
    colunas = pd.concat(args, axis=1)
    matriz_colunas = np.array(colunas)
    lista_concatenada = []
    for a in range(len(matriz_colunas)):
        if 'x' in list(matriz_colunas[a]):
            lista_concatenada.append('x')
        else:
            lista_concatenada.append(None)

    df_concatenado = pd.DataFrame(lista_concatenada)
    df_concatenado.columns = coluna
    return df_concatenado


def PadraoLVC(df):
    def RandonAreas(df):
        # Calcula área trincada pela metodologia antiga - FC-2 / FC-3 / AT/ON / P
        data = np.random.randint(11, 66, size=(df.shape[0], 9))
        data_p = np.random.randint(1, 8, size=df.shape[0])
        df_new = pd.DataFrame(data, columns=['FC-2 J','FC-2 TB', 'FC-3 JE', 'FC-3 TBE', 'AT/ON 1', 'AT/ON 2', 'AT/ON 3', 'AT/ON 4', 'AT/ON 5'])
        df_new2 = pd.DataFrame(data_p, columns=['PA'])
        df = pd.concat([df, df_new, df_new2], axis = 1)
        df['FC-2 J'].where(df['J'] == "x", 0, inplace=True)
        df['FC-2 TB'].where(df['TB'] == "x", 0, inplace=True)
        df['FC-2'] = df['FC-2 J'] + df['FC-2 TB']
        df['FC-3 JE'].where(df['JE'] == "x", 0, inplace=True)
        df['FC-3 TBE'].where(df['TBE'] == "x", 0, inplace=True)
        df['FC-3'] = df['FC-3 JE'] + df['FC-3 TBE']
        df['AT/ON 1'].where(df['ALP'] == "x", 0, inplace=True)
        df['AT/ON 2'].where(df['ALC'] == "x", 0, inplace=True)
        df['AT/ON 3'].where(df['ATP'] == "x", 0, inplace=True)
        df['AT/ON 4'].where(df['ATC'] == "x", 0, inplace=True)
        df['AT/ON 5'].where(df['O'] == "x", 0, inplace=True)
        df['AT/ON'] = df['AT/ON 1'] + df['AT/ON 2'] + df['AT/ON 3'] + df['AT/ON 4'] + df['AT/ON 5']
        df['PA'].where(df['P'] == "x", 0, inplace=True)
        return df
    
    # Recebe varios df de mesmo tamanho e concatena os pontos onde foram detectadas patologias na mesma linha
    columns_names_list = list(df.columns.values.tolist())
    for name in columns_names_list:
        if "ATRE" in name:
            teclas_var = True
            break
        else:
            teclas_var = False
    
    if teclas_var == True:
        df_LVC = pd.DataFrame()
        df_ini_fim = pd.concat([df['Início'], df['Fim']], axis=1)
        df_ini_fim.columns = ['Início', 'Fim']
        df_TRR = ConcatenaDf(['TRR'], df['TRR'])
        df_O = ConcatenaDf(['O'], df['OND.BE'], df['OND.ATRE'], df['OND.F'], df['OND.ATRD'], df['OND.BD'])
        df_P = ConcatenaDf(['P'], df['Panela.BE.A'], df['Panela.BE.M'], df['Panela.BE.B'], df['Panela.ATRE.A'], df['Panela.ATRE.M'], df['Panela.ATRBE.B'], df['Panela.F.A'], 
        df['Panela.F.M'], df['Panela.F.B'], df['Panela.ATRD.A'], df['Panela.ATRD.M'], df['Panela.ATRD.B'], df['Panela.BD.A'], df['Panela.BD.M'], df['Panela.BD.B'])
        df_E = ConcatenaDf(['E'], df['E'])
        df_Ex = ConcatenaDf(['Ex'], df['Exsudação.BE'], df['Exsudação.ATRE'], df['Exsudação.F'], df['Exsudação.ATRD'], df['Exsudação.BD'])
        df_D = ConcatenaDf(['D'], df['D'])
        df_R = ConcatenaDf(['R'], df['Remendo.BE'], df['Remendo.ATRE'], df['Remendo.F'], df['Remendo.ATRD'], df['Remendo.BD'])
        df_Fi = ConcatenaDf(['Fi'], df['Fi.FC-1.BE'], df['Fi.FC-1.ATRE'], df['Fi.FC-1.F'], df['Fi.FC-1.ATRD'], df['Fi.FC-1.BD'], df['J1.FC-1.BE'], df['J1.FC-1.ATRE'], df['J1.FC-1.F'], df['J1.FC-1.ATRD'], df['J1.FC-1.BD'])
        df_J = ConcatenaDf(['J'], df['J.FC-2.BE'], df['J.FC-2.ATRE'], df['J.FC-2.F'], df['J.FC-2.ATRD'], df['J.FC-2.BD'])
        df_JE = ConcatenaDf(['JE'], df['JE.FC-3.BE'], df['JE.FC-3.ATRE'], df['JE.FC-3.F'], df['JE.FC-3.ATRD'], df['JE.FC-3.BD'])
        df_TB = ConcatenaDf(['TB'], df['TB'])
        df_TBE = ConcatenaDf(['TBE'], df['TBE'])
        df_TTC = ConcatenaDf(['TTC'], df['TTC.FC-23.BE'], df['TTC.FC-23.ATRE'], df['TTC.FC-23.F'], df['TTC.FC-23.ATRD'], df['TTC.FC-23.BD'])
        df_TTL = ConcatenaDf(['TTL'], df['TTL.FC-23.BE'], df['TTL.FC-23.ATRE'], df['TTL.FC-23.F'], df['TTL.FC-23.ATRD'], df['TTL.FC-23.BD'])
        df_TLC = ConcatenaDf(['TLC'], df['TLC.FC-23.BE'], df['TLC.FC-23.ATRE'], df['TLC.FC-23.F'], df['TLC.FC-23.ATRD'], df['TLC.FC-23.BD'])
        df_TLL = ConcatenaDf(['TLL'], df['TLL.FC-23.BE'], df['TLL.FC-23.ATRE'], df['TLL.FC-23.F'], df['TLL.FC-23.ATRD'], df['TLL.FC-23.BD'])
        df_ALP = ConcatenaDf(['ALP'], df['ALP-23.BE'], df['ALP-23.ATRE'], df['ALP-23.F'], df['ALP-23.ATRD'], df['ALP-23.BD'])
        df_ALC = ConcatenaDf(['ALC'], df['ALC-23.BE'], df['ALC-23.ATRE'], df['ALC-23.F'], df['ALC-23.ATRD'], df['ALC-23.BD'])
        df_ATP = ConcatenaDf(['ATP'], df['ATP-23.BE'], df['ATP-23.ATRE'], df['ATP-23.F'], df['ATP-23.ATRD'], df['ATP-23.BD'])
        df_ATC = ConcatenaDf(['ATC'], df['ATC-23.BE'], df['ATC-23.ATRE'], df['ATC-23.F'], df['ATC-23.ATRD'], df['ATC-23.BD'])
        try:
            df_dg_hora = pd.concat([df['DG'], df['Observação'], df['Latitude'], df['Longitude'], df['Altitude'], df['Data'], df['Hora']], axis=1)
            df_dg_hora.columns = ['DG', 'Observação', 'Latitude', 'Longitude', 'Altitude', 'Data', 'Hora']
        except:
            df_dg_hora = pd.concat([df['DG'], df['Observação'], df['Latitude'], df['Longitude'], df['Altitude']], axis=1)
            df_dg_hora.columns = ['DG', 'Observação', 'Latitude', 'Longitude', 'Altitude']
        
        df_LVC = pd.concat([df_ini_fim, df_TRR, df_O, df_P, df_E, df_Ex, df_D, df_R, df_Fi, df_J, df_JE, df_TB, df_TBE, df_TTC, df_TTL, df_TLC, df_TLL, df_ALP, df_ALC, df_ATP, df_ATC, df_dg_hora], axis=1)
        df_LVC = RandonAreas(df_LVC)
        return df_LVC, teclas_var
    
    elif teclas_var == False:
        df = RandonAreas(df)
        return df, teclas_var


def InverteDecrescente(df):
    # Função morta
    df = df.sort_values(by = 'Início', ascending=True)
    first_column = df.pop('Fim')
    df.insert(0, 'Fim', first_column)
    df.rename(columns = {'Início': 'Fim', 'Fim': 'Início'}, inplace=True)
    return df


def CriaEstacoes(km, passo):
    # Recebe uma lista de kms e separa aqueles dentro do espassamento desejado
    estacoes = []
    passo = round(passo, 3)
    for line in km:
        if round(round(line*1000, 3) % (passo*1000)) == 0:
            estacoes.append(round(line,4))
    return estacoes


def FiltroDeTrincas(df_LVC, estacao):
    # ANOTA ONDE NA LINHA AVALIADA TEM JE, TBE, J, TB
    indice = df_LVC.loc[round(df_LVC['Início']*1000, 3) == estacao].index.tolist()[0]
    vetor_JE, vetor_TBE, vetor_J, vetor_TB, vetor_isoladas = False, False, False, False, False
    if df_LVC.loc[indice, 'JE'] == 'x':
        vetor_JE = True
    elif df_LVC.loc[indice, 'TBE'] == 'x':
        vetor_TBE = True
    elif df_LVC.loc[indice, 'J'] == 'x':
        vetor_J = True
    elif df_LVC.loc[indice, 'TB'] == 'x':
        vetor_TB = True
    elif (df_LVC.loc[indice, 'TTC'] == 'x') or (df_LVC.loc[indice, 'TTL'] == 'x') or (df_LVC.loc[indice, 'TLC'] == 'x') or (df_LVC.loc[indice, 'TLL'] == 'x') or (df_LVC.loc[indice, 'TRR'] == 'x'):
        vetor_isoladas = True
    return [indice, vetor_JE, vetor_TBE, vetor_J, vetor_TB, vetor_isoladas]


def FiltroEstacoes(estacoes, df_LVC, df_ATR, df_area, crescente, artesp, teclas_var):
    # DEIXA APENAS OS DADOS DAS ESTAÇÕES E LIMPA EXCESSO DE TRINCAS
    # Calcula área trincada quando necessário | Concatena observações | Cria coluna OK
    df_filtrado = pd.DataFrame()
    df_ocorencia_absoluta = pd.DataFrame()
    for estacao in estacoes:
        df_temp = pd.DataFrame()
        df_temp_absoluto = pd.DataFrame()
        pontos_trincados = []
        # Pega dados da estação
        df_temp = pd.concat([df_temp, df_LVC.loc[round(df_LVC['Início']*1000, 3) == round(estacao*1000, 3)]])
        pontos_trincados.append(FiltroDeTrincas(df_LVC, round(estacao*1000, 3)))
        # Pega ATR da estação
        indice_atr = df_ATR.loc[round(df_ATR['Inicio (km)']*1000, 3) == round(estacao*1000, 3)].index
        df_temp['ATR Esq (mm)'] = df_ATR.iloc[indice_atr[0]]['ATR Esq (mm)']
        df_temp['ATR Dir (mm)'] = df_ATR.iloc[indice_atr[0]]['ATR Dir (mm)']
        # Coluna OK-ATR para as versões ANTT
        df_temp.insert(2, 'OK-ATR', None, True)
        if (df_temp['ATR Esq (mm)'].max() > 7) or (df_temp['ATR Dir (mm)'].max() > 7):
            pass
        else:
            df_temp['OK-ATR'] = 'x'

        # Pega restante das linhas se existirem - VERSÃO 1m
        if abs(round((df_temp['Início'].tolist()[0] - df_temp['Fim'].tolist()[0])*1000, 3)) < 2:
            try:
                df_temp = pd.concat([df_temp, df_LVC.loc[round(df_LVC['Início']*1000, 3) == round(((estacao*1000) - 1), 3)]])
                pontos_trincados.append(FiltroDeTrincas(df_LVC, round(((estacao*1000) - 1), 3)))
            except:
                pass
            try:
                df_temp = pd.concat([df_temp, df_LVC.loc[round(df_LVC['Início']*1000, 3) == round(((estacao*1000) + 1), 3)]])
                pontos_trincados.append(FiltroDeTrincas(df_LVC, round(((estacao*1000) + 1), 3)))
            except:
                pass
            try:
                df_temp = pd.concat([df_temp, df_LVC.loc[round(df_LVC['Início']*1000, 3) == round(((estacao*1000) - 2), 3)]])
                pontos_trincados.append(FiltroDeTrincas(df_LVC, round(((estacao*1000) - 2), 3)))
            except:
                pass
            try:
                df_temp = pd.concat([df_temp, df_LVC.loc[round(df_LVC['Início']*1000, 3) == round(((estacao*1000) + 2), 3)]])
                pontos_trincados.append(FiltroDeTrincas(df_LVC, round(((estacao*1000) + 2), 3)))
            except:
                pass
            if crescente:
                try:
                    df_temp = pd.concat([df_temp, df_LVC.loc[round(df_LVC['Início']*1000, 3) == round(((estacao*1000) + 3), 3)]])
                    pontos_trincados.append(FiltroDeTrincas(df_LVC, round(((estacao*1000) + 3), 3)))
                except:
                    pass
            if not crescente:
                try:
                    df_temp = pd.concat([df_temp, df_LVC.loc[round(df_LVC['Início']*1000, 3) == round(((estacao*1000) - 3), 3)]])
                    pontos_trincados.append(FiltroDeTrincas(df_LVC, round(((estacao*1000) - 3), 3)))
                except:
                    pass
        # Pega restante das linhas se existirem - VERSÃO 2m
        elif abs(round((df_temp['Início'].tolist()[0] - df_temp['Fim'].tolist()[0])*1000, 3)) == 2:
            try:
                df_temp = pd.concat([df_temp, df_LVC.loc[round(df_LVC['Início']*1000, 3) == round(((estacao*1000) - 1), 3)]])
                pontos_trincados.append(FiltroDeTrincas(df_LVC, round(((estacao*1000) - 1), 3)))
            except:
                pass
            try:
                df_temp = pd.concat([df_temp, df_LVC.loc[round(df_LVC['Início']*1000, 3) == round(((estacao*1000) + 1), 3)]])
                pontos_trincados.append(FiltroDeTrincas(df_LVC, round(((estacao*1000) + 1), 3)))
            except:
                pass

        # Concatenando OBS das estações
        obs=""
        if crescente:
            df_obs = df_LVC[(df_LVC['Início']*1000 >= round(((estacao*1000) - 10), 3)) & (df_LVC['Início']*1000 < round(((estacao*1000) + 10), 3))]
        elif not crescente:
            df_obs = df_LVC[(df_LVC['Início']*1000 <= round(((estacao*1000) + 10), 3)) & (df_LVC['Início']*1000 > round(((estacao*1000) - 10), 3))]
        for observacao in df_obs['Observação']:
            if obs and not pd.isna(observacao):
                obs=obs+" | "+str(observacao)
            elif not pd.isna(observacao):
                obs=str(observacao)

        # Pega área das patologias necessárias para a P15
        if teclas_var and artesp:
            df_temp['FC-2'] = 0
            df_temp['FC-3'] = 0
            df_temp['PA'] = 0
            df_temp['AT/ON'] = 0
            if crescente:
                df_area_estacao = df_area[(df_area['Inicial']*1000 >= round(((estacao*1000) - 10), 3)) & (df_area['Final']*1000 <= round(((estacao*1000) + 10), 3))]
            elif not crescente:
                df_area_estacao = df_area[(df_area['Inicial']*1000 <= round(((estacao*1000) + 10), 3)) & (df_area['Final']*1000 >= round(((estacao*1000) - 10), 3))]
            df_area_j2 = df_area_estacao[(df_area_estacao['Sigla'] == 'J')]
            df_area_j3 = df_area_estacao[(df_area_estacao['Sigla'] == 'JE')]
            df_area_pa = df_area_estacao[(df_area_estacao['Sigla'] == 'P')]
            df_area_afu = df_area_estacao[(df_area_estacao['Sigla'] == 'Afund/Ond')]
            # Eliminando áreas de defeitos em pontos sem marcação na estação
            if 'x' in df_temp['J'].tolist():
                df_temp['FC-2'] = (df_area_j2['Área (m²)'].sum())/(3.6*20)
            if 'x' in df_temp['TB'].tolist() and 'x' not in df_temp['J'].tolist():
                df_temp['FC-2'] = randint(10,40)/100
            if 'x' in df_temp['JE'].tolist():
                df_temp['FC-3'] = (df_area_j3['Área (m²)'].sum())/(3.6*20)
            if 'x' in df_temp['TBE'].tolist() and 'x' not in df_temp['JE'].tolist():
                df_temp['FC-3'] = randint(10,40)/100
            if 'x' in df_temp['P'].tolist():
                df_temp['PA'] = (df_area_pa['Área (m²)'].sum())/(3.6*20)
            for rut in ['ALP', 'ALC', 'ATP', 'ATC', 'O']:
                if 'x' in df_temp[rut].tolist():
                    df_temp['AT/ON'] = (df_area_afu['Área (m²)'].sum())/(3.6*20)

        # Cálculo da média de FC-2 / FC-3 / AT/ON / P na estação - Aproveitando o df_obs que já esta no intervalo desejado
        else:
            df_temp['FC-2'] = (df_obs['FC-2'].mean())/100
            df_temp['FC-3'] = (df_obs['FC-3'].mean())/100
            df_temp['PA'] = (df_obs['PA'].mean())/100
            df_temp['AT/ON'] = (df_obs['AT/ON'].mean())/100
            # Eliminando áreas de defeitos em pontos sem marcação na estação
            if ('x' not in df_temp['TB'].tolist()) and ('x' not in df_temp['J'].tolist()):
                df_temp['FC-2'] = 0
            if ('x' not in df_temp['TBE'].tolist()) and ('x' not in df_temp['JE'].tolist()):
                df_temp['FC-3'] = 0
            if 'x' not in df_temp['P'].tolist():
                df_temp['PA'] = 0
            if ('x' not in df_temp['ALP'].tolist()) and ('x' not in df_temp['ALC'].tolist()) and ('x' not in df_temp['ATP'].tolist()) and ('x' not in df_temp['ATC'].tolist()) and ('x' not in df_temp['O'].tolist()):
                df_temp['AT/ON'] = 0

        df_temp_absoluto = df_temp.copy(deep=True)
        for metro in pontos_trincados: # limpando excesso de trincas por estação --- JE_TBE_J_TB
            if metro[1] is True:
                df_temp['TBE'] = df_temp['TBE'].replace(['x'], None)
                df_temp['J'] = df_temp['J'].replace(['x'], None)
                df_temp['TB'] = df_temp['TB'].replace(['x'], None)
                df_temp['TTC'] = df_temp['TTC'].replace(['x'], None)
                df_temp['TTL'] = df_temp['TTL'].replace(['x'], None)
                df_temp['TLC'] = df_temp['TLC'].replace(['x'], None)
                df_temp['TLL'] = df_temp['TLL'].replace(['x'], None)
                df_temp['TRR'] = df_temp['TRR'].replace(['x'], None)
                df_temp['Fi'] = df_temp['Fi'].replace(['x'], None)
            elif metro[2] is True:
                df_temp['J'] = df_temp['J'].replace(['x'], None)
                df_temp['TB'] = df_temp['TB'].replace(['x'], None)
                df_temp['TTC'] = df_temp['TTC'].replace(['x'], None)
                df_temp['TTL'] = df_temp['TTL'].replace(['x'], None)
                df_temp['TLC'] = df_temp['TLC'].replace(['x'], None)
                df_temp['TLL'] = df_temp['TLL'].replace(['x'], None)
                df_temp['TRR'] = df_temp['TRR'].replace(['x'], None)
                df_temp['Fi'] = df_temp['Fi'].replace(['x'], None)
            elif metro[3] is True:
                df_temp['TB'] = df_temp['TB'].replace(['x'], None)
                df_temp['TTC'] = df_temp['TTC'].replace(['x'], None)
                df_temp['TTL'] = df_temp['TTL'].replace(['x'], None)
                df_temp['TLC'] = df_temp['TLC'].replace(['x'], None)
                df_temp['TLL'] = df_temp['TLL'].replace(['x'], None)
                df_temp['TRR'] = df_temp['TRR'].replace(['x'], None)
                df_temp['Fi'] = df_temp['Fi'].replace(['x'], None)
            elif metro[4] is True:
                df_temp['TTC'] = df_temp['TTC'].replace(['x'], None)
                df_temp['TTL'] = df_temp['TTL'].replace(['x'], None)
                df_temp['TLC'] = df_temp['TLC'].replace(['x'], None)
                df_temp['TLL'] = df_temp['TLL'].replace(['x'], None)
                df_temp['TRR'] = df_temp['TRR'].replace(['x'], None)
                df_temp['Fi'] = df_temp['Fi'].replace(['x'], None)
            elif metro[5] is True:
                df_temp['TTC'] = df_temp['TTC'].replace(['x'], None)
                df_temp['TTL'] = df_temp['TTL'].replace(['x'], None)
                df_temp['TLC'] = df_temp['TLC'].replace(['x'], None)
                df_temp['TLL'] = df_temp['TLL'].replace(['x'], None)
                df_temp['TRR'] = df_temp['TRR'].replace(['x'], None)
                df_temp['Fi'] = 'x'

        # Deixando apenas as estações
        colunas_LVC = ['TRR', 'O', 'P', 'E', 'Ex', 'D', 'R', 'Fi', 'J', 'JE', 'TB', 'TBE', 'TTC', 'TTL', 'TLC', 'TLL', 'ALP', 'ALC', 'ATP', 'ATC']
        coluna_ok = True
        for colunas in colunas_LVC:
            if 'x' in df_temp[colunas].tolist():
                df_temp[colunas] = 'x'
                coluna_ok = False
            else:
                df_temp[colunas] = None
            if 'x' in df_temp_absoluto[colunas].tolist():
                df_temp_absoluto[colunas] = 'x'
            else:
                df_temp_absoluto[colunas] = None

        for linha in range(len(pontos_trincados) - 1):
            df_temp.drop([pontos_trincados[linha + 1][0]], inplace=True)
            df_temp_absoluto.drop([pontos_trincados[linha + 1][0]], inplace=True)
        df_temp.iat[0, 23] = obs
        df_temp_absoluto.iat[0, 23] = obs
        # Ajustado coluna OK
        df_temp.insert(2, 'OK', None, True)
        df_temp_absoluto.insert(2, 'OK', None, True)
        if coluna_ok == True:
            df_temp['OK'] = 'x'
            df_temp_absoluto['OK'] = 'x'
        else:
            df_temp['OK'] = None
            df_temp_absoluto['OK'] = None
        
        # Concatenando df final com o os demais
        df_filtrado = pd.concat([df_filtrado, df_temp], ignore_index=True)
        df_ocorencia_absoluta = pd.concat([df_ocorencia_absoluta, df_temp_absoluto], ignore_index=True)

    return df_filtrado, df_ocorencia_absoluta


def UnirPistaSimples(df_crescente_filtrado, df_decrescente_filtrado, estacoes_crescente): # Intercala estações das pistas simples
    df_concat = pd.DataFrame()
    df_concat_crescente = pd.DataFrame()
    df_concat_decrescente = pd.DataFrame()
    estacao = min(estacoes_crescente)
    while estacao < max(estacoes_crescente):
        df_concat = pd.concat([df_concat, df_crescente_filtrado.loc[df_crescente_filtrado['Início']*1000 == estacao*1000]], ignore_index = True)
        df_concat_crescente = pd.concat([df_concat_crescente, df_crescente_filtrado.loc[df_crescente_filtrado['Início']*1000 == estacao*1000]], ignore_index = True)
        estacao = round((estacao*1000) + (0.02*1000))/1000
        try:
            df_concat = pd.concat([df_concat, df_decrescente_filtrado.loc[df_decrescente_filtrado['Início']*1000 == estacao*1000]], ignore_index = True)
            df_concat_decrescente = pd.concat([df_concat_decrescente, df_decrescente_filtrado.loc[df_decrescente_filtrado['Início']*1000 == estacao*1000]], ignore_index = True)
        except:
            pass
        estacao = round((estacao*1000) + (0.02*1000))/1000
    df_concat_decrescente.sort_values(by=['Início'], ascending=False, inplace=True, ignore_index=True)
    return df_concat, df_concat_crescente, df_concat_decrescente


def FichaIGG(df, df_abs_concat, num_estacoes, ini_fim):
    # Calcula todos os dados necessários na Ficha IGG em todos os modelos de apresentação. 
    ocorrencias = {'TRR':0, 'O':0, 'P':0, 'E':0, 'Ex':0, 'D':0, 'R':0, 'Fi':0, 'J':0, 'JE':0, 'TB':0, 'TBE':0, 'TTC':0, 'TTL':0, 'TLC':0, 'TLL':0, 'ALP':0, 'ALC':0, 'ATP':0, 'ATC':0}
    ocorrencias_abs = {'TRR':0, 'O':0, 'P':0, 'E':0, 'Ex':0, 'D':0, 'R':0, 'Fi':0, 'J':0, 'JE':0, 'TB':0, 'TBE':0, 'TTC':0, 'TTL':0, 'TLC':0, 'TLL':0, 'ALP':0, 'ALC':0, 'ATP':0, 'ATC':0}
    for coluna in ocorrencias:
        ocorrencias[coluna] = (df[df[coluna] == "x"][coluna].count())
        ocorrencias_abs[coluna] = (df_abs_concat[df_abs_concat[coluna] == "x"][coluna].count())

    trincas_isoladas = ocorrencias['Fi'] + ocorrencias['TTC'] + ocorrencias['TTL'] + ocorrencias['TLC'] + ocorrencias['TLL'] + ocorrencias['TRR']
    FC_2 = ocorrencias['J'] + ocorrencias['TB']
    FC_3 = ocorrencias['JE'] + ocorrencias['TBE']
    afundamentos = ocorrencias['ALP'] + ocorrencias['ATP'] + ocorrencias['ALC'] + ocorrencias['ATC']
    desagrecacao = ocorrencias['O'] + ocorrencias['P'] + ocorrencias['E']
    exudacao = ocorrencias['Ex']
    desgaste = ocorrencias['D']
    remendos = ocorrencias['R']

    trincas_isoladas_abs = ocorrencias_abs['Fi'] + ocorrencias_abs['TTC'] + ocorrencias_abs['TTL'] + ocorrencias_abs['TLC'] + ocorrencias_abs['TLL'] + ocorrencias_abs['TRR']
    FC_2_abs = ocorrencias_abs['J'] + ocorrencias_abs['TB']
    FC_3_abs = ocorrencias_abs['JE'] + ocorrencias_abs['TBE']

    fa = [trincas_isoladas_abs, FC_2_abs, FC_3_abs, afundamentos, desagrecacao, exudacao, desgaste, remendos]
    fac = [trincas_isoladas, FC_2, FC_3]

    fr = [trincas_isoladas * 100 / num_estacoes, 
          FC_2 * 100 / num_estacoes, 
          FC_3 * 100 / num_estacoes, 
          afundamentos * 100 / num_estacoes, 
          desagrecacao * 100 / num_estacoes, 
          exudacao * 100 / num_estacoes, 
          desgaste * 100 / num_estacoes, 
          remendos * 100 / num_estacoes]

    med_esq = df['ATR Esq (mm)'].mean()
    med_dir = df['ATR Dir (mm)'].mean()
    desv_esq = df['ATR Esq (mm)'].std()
    desv_dir = df['ATR Dir (mm)'].std()

    # Médias e Observações para P21
    fc2 = df['FC-2'].mean()
    fc3 = df['FC-3'].mean()

    # Verificar como tirar esse try
    km_ref = ""
    ramo = ""
    faixa = ""
    try:
        km_ref = df["km referencia"].tolist()[0]
        ramo = df["Nome ramo"].tolist()[0]
        faixa = df["Faixa"].tolist()[0]
    except:
        pass
    
    # Defesa para caso não tenha valores validos
    for tipo in range(len(fr)):
        if math.isnan(fr[tipo]):
            fr[tipo] = 0
    if math.isnan(desv_esq):
        desv_esq = 0
    if math.isnan(desv_dir):
        desv_dir = 0

    if math.isnan(med_esq) or med_esq == "-":
        med_esq = 0
    if math.isnan(med_dir) or med_dir == "-":
        med_dir = 0

    var_esq = (desv_esq)**2
    var_dir = (desv_dir)**2

    # ajuste dos valores de atr
    if med_esq <= 30:
        med_esq_igi = med_esq * (4/3)
    else:
        med_esq_igi = 40
    if med_dir <= 30:
        med_dir_igi = med_dir * (4/3)
    else:
        med_dir_igi = 40
    # ajuste dos valores de variancia
    if var_esq > 50:
        var_esq_igi = 50
    else:
        var_esq_igi = var_esq
    if var_dir > 50:
        var_dir_igi = 50
    else:
        var_dir_igi = var_dir

    igi = [fr[0] * 0.2,
           fr[1] * 0.5, 
           fr[2] * 0.8,
           fr[3] * 0.9,
           fr[4] * 1,
           fr[5] * 0.5,
           fr[6] * 0.3,
           fr[7] * 0.6,
           (med_esq_igi + med_dir_igi)/2,
           (var_esq_igi + var_dir_igi)/2]

    igg = sum(igi)
    if igg <= 20:
        conceito = 'Ótimo'
    elif igg > 20 and igg <= 40:
        conceito = 'Bom'
    elif igg > 40 and igg <= 80:
        conceito = 'Regular'
    elif igg > 80 and igg <= 160:
        conceito = 'Ruim'
    elif igg > 160:
        conceito = 'Péssimo'
    return {"Início": ini_fim[0], "Fim": ini_fim[-1], "Freq. Absoluta": fa, 
            "Freq. Abs. Consi.": fac, "Freq. Relativa": fr, "IGI": igi, "IGG": igg, "Med. Esq.": med_esq, 
            "Med. Dir.": med_dir, "Desv. Esq.": desv_esq, "Desv. Dir.": desv_dir, "Var. Esq.": var_esq, 
            "Var. Dir.": var_dir, "Conceito": conceito, "Num. Estacoes": num_estacoes,
            "FC-2": fc2, "FC-3": fc3, "km referencia": km_ref, "Nome ramo": ramo, "Faixa": faixa}


def ColunasFaltantes(df, cabecalho):
    # Cria a divisão dos km por estacas de 20m
    estaca=[]
    resto_estaca=[]
    for linha in df["Início"]:
        estaca.append(round((linha*1000)/20, 0))
        resto_estaca.append(round(((linha*1000)%20), 3))
    # Adiciona colunas faltantes no df | Necessário para a P15
    df.insert(0, "Estaca", estaca, True)
    df.insert(1, "+", '+', True)
    df.insert(2, "Resto estaca", resto_estaca, True)
    df.insert(0, "Tipo Revest.", cabecalho["Tipo Revest."], True)
    df.insert(0, "Secao Terra.", cabecalho["Secao Terra."], True)
    df.insert(0, "Faixa", cabecalho["Faixa"], True)
    df.insert(0, "Nome ramo", cabecalho["Nome ramo"], True)
    df.insert(0, "km referencia", cabecalho["km referencia"], True)
    df.insert(0, "Sentido", cabecalho["Sentido"].upper(), True)
    df.insert(0, "Lado", cabecalho["Lado"], True)
    df.insert(0, "Tipo pista", cabecalho["Tipo pista"].upper(), True)
    df.insert(0, "Via", cabecalho["Via"].upper(), True)
    df.insert(0, "Rodovia", cabecalho["Rodovia"], True)
    df.insert(0, "Lote", cabecalho["Lote"], True)
    df.insert(0, "STH", cabecalho["STH"], True)
    return df


def GeraP15(main_path, modelo_path, file, planilha):
    # Ajustando x para X
    for column in ['OK', 'Fi', 'TTC', 'TTL', 'TLC', 'TLL', 'TRR', 'J', 'TB', 'JE', 'TBE', 'ALP', 'ATP', 'O', 'P', 'Ex', 'D', 'R', 'ALC', 'ATC', 'E']:
        planilha[column].where(planilha[column] != "x", "X", inplace=True)
    for atr in ['ATR Dir (mm)', 'ATR Esq (mm)']:
        planilha[atr].where(planilha[atr] != None, "-", inplace=True)

    # Abrindo arquivo modelo
    print("Iniciando exportação do arquivo " + file)
    book = op.load_workbook(modelo_path)

    # Editando arquivo
    sheet = book.active
    sheet.title = '_'
    sheet = book['_']
    start_col = 1
    start_row = 4
    end_col = 44

    # Copia formatação
    for line in range(planilha.shape[0]):
        for col in range(end_col):
            sheet.cell(row = start_row + line, column = start_col + col)._style = copy(sheet.cell(row = start_row, column = col+1)._style)
    
    anexo_colunas = ['Lote', 'Rodovia', 'Via', 'Tipo pista', 'Sentido', 'km referencia', 'Nome ramo', 'Faixa', 'Início', 'Secao Terra.', 'Tipo Revest.', 'OK', 'Fi', 'TTC', 'TTL', 'TLC', 'TLL', 'TRR',
                      'J', 'TB', 'JE', 'TBE', 'ALP', 'ATP', 'O', 'P', 'Ex', 'D', 'R', 'ALC', 'ATC', 'E', 'ATR Dir (mm)', 'ATR Esq (mm)', 'FC-2', 'FC-3', 'PA', 'AT/ON', 'Data', 'Hora', 'Observação', 'Latitude', 'Longitude', 'Altitude']
    for line in tqdm(range(planilha.shape[0])):
        for col in range(len(anexo_colunas)):
            sheet.cell(row = start_row + line, column = start_col + col).value = planilha[anexo_colunas[col]][line]
            sheet.row_dimensions[start_row + line].height = 15.5

    sheet.print_area = 'A1:AR' + str(planilha.shape[0] + start_row - 1)
    nomeIGG = os.path.join(main_path, file)
    book.save(nomeIGG)
    print("    ... pronto para uso \n")


def AnexoAPavesys(main_path, modelo_path, file, cabecalho, planilha2):
    # Problema do X que não conta na versão pavesys_faixa - Este improviso deve ser revisado no futuro
    planilha = planilha2.copy(deep=True)

    # Ajustando x para X
    for column in ['OK', 'Fi', 'TTC', 'TTL', 'TLC', 'TLL', 'TRR', 'J', 'TB', 'JE', 'TBE', 'ALP', 'ATP', 'ALC', 'ATC', 'O', 'P', 'E', 'Ex', 'D', 'R']:
        planilha[column].where(planilha[column] != "x", "X", inplace=True)
    for atr in ['ATR Dir (mm)', 'ATR Esq (mm)']:
        planilha[atr].where(planilha[atr] != None, "-", inplace=True)

    # Abrindo arquivo modelo
    new_name = (file.split("PP_")[1]).split(".")[0] + ".xlsx"
    print("Iniciando exportação do arquivo: Anexo A - Pavesys - " + new_name)
    nomeIGG = os.path.join(main_path, 'Anexo A - Pavesys - ' + new_name)
    shutil.copy(modelo_path, nomeIGG)
    book = op.load_workbook(nomeIGG)

    # Editando arquivo
    sheet = book["_"]
    start_col = 1
    start_row = 10
    end_col = 26

    # Copia formatação
    for line in range(planilha.shape[0]):
        for col in range(end_col):
            sheet.cell(row = start_row + line, column = start_col + col)._style = copy(sheet.cell(row = start_row, column = col+1)._style)
            sheet.row_dimensions[start_row + line].height = 17.5

    anexo_colunas = ['Início', 'Lado', 'OK', 'Fi', 'TTC', 'TTL', 'TLC', 'TLL', 'TRR', 'J', 'TB', 'JE', 'TBE', 'ALP', 'ATP', 'ALC', 'ATC', 'O', 'P', 'E', 'Ex', 'D', 'R', 'ATR Esq (mm)', 'ATR Dir (mm)', 'Observação']
    for line in range(planilha.shape[0]):
        for col in range(len(anexo_colunas)):
            sheet.cell(row = start_row + line, column = start_col + col).value = planilha[anexo_colunas[col]][line]
        
    sheet.cell(row = 2, column = 1).value = "RODOVIA: " + cabecalho["Rodovia"]
    sheet.cell(row = 3, column = 1).value = "TRECHO: " + cabecalho["STH"]
    sheet.cell(row = 4, column = 1).value = "PISTA: " + cabecalho["Tipo pista"]
    sheet.cell(row = 4, column = 6).value = "SENTIDO: " + cabecalho["Sentido"]
    sheet.cell(row = 4, column = 10).value = "FAIXA: " + cabecalho["Faixa"]
    sheet.cell(row = 2, column = 14).value = "OPERADOR: " + cabecalho["Operador"]
    sheet.cell(row = 3, column = 14).value = "REVESTIMENTO TIPO: " + cabecalho["Tipo Revest."]
    sheet.cell(row = 4, column = 14).value = "DATA: " + cabecalho["Data"]
    sheet.cell(row = 4, column = 23).value = cabecalho["km ini"]*1000
    sheet.cell(row = 4, column = 26).value = cabecalho["km fim"]*1000

    sheet.print_area = 'A1:Z' + str(planilha.shape[0] + start_row - 1)
    book.save(nomeIGG)
    print("    ... pronto para uso \n")


def AnexoADERSP(main_path, modelo_path, file, cabecalho, planilha):
    # Ajustando x para X
    for column in ['OK', 'Fi', 'TTC', 'TTL', 'TLC', 'TLL', 'TRR', 'J', 'TB', 'JE', 'TBE', 'ALP', 'ATP', 'ALC', 'ATC', 'O', 'P', 'E', 'Ex', 'D', 'R']:
        planilha[column].where(planilha[column] != "x", "X", inplace=True)
    for atr in ['ATR Dir (mm)', 'ATR Esq (mm)']:
        planilha[atr].where(planilha[atr] != None, "-", inplace=True)

    # Abrindo arquivo modelo
    new_name = (file.split("PP_")[1]).split(".")[0] + ".xlsx"
    print("Iniciando exportação do arquivo: Anexo A - DER-SP - " + new_name)
    nomeIGG = os.path.join(main_path, 'Anexo A - DER-SP - ' + new_name)
    shutil.copy(modelo_path, nomeIGG)
    book = op.load_workbook(nomeIGG)

    # Editando arquivo
    sheet = book.active
    sheet.title = '_'
    sheet = book['_']
    start_col = 1
    start_row = 10
    end_col = 29

    # Copia formatação
    for line in range(planilha.shape[0]):
        for col in range(end_col):
            sheet.cell(row = start_row + line, column = start_col + col)._style = copy(sheet.cell(row = start_row, column = col+1)._style)
            sheet.row_dimensions[start_row + line].height = 17.5

    anexo_colunas = ['Estaca', '+', 'Resto estaca', 'Início', 'Lado', 'OK', 'Fi', 'TTC', 'TTL', 'TLC', 'TLL', 'TRR', 'J', 'TB', 'JE', 'TBE', 'ALP', 'ATP', 'ALC', 'ATC', 'O', 'P', 'E', 'Ex', 'D', 'R', 'ATR Esq (mm)', 'ATR Dir (mm)', 'Observação']
    for line in range(planilha.shape[0]):
        for col in range(len(anexo_colunas)):
            sheet.cell(row = start_row + line, column = start_col + col).value = planilha[anexo_colunas[col]][line]
        
    sheet.cell(row = 2, column = 1).value = "RODOVIA: " + cabecalho["Rodovia"]
    sheet.cell(row = 3, column = 1).value = "TRECHO: " + cabecalho["Trecho"]
    sheet.cell(row = 4, column = 1).value = "PISTA: " + cabecalho["Tipo pista"]
    sheet.cell(row = 2, column = 17).value = "OPERADOR: " + cabecalho["Operador"]
    sheet.cell(row = 3, column = 17).value = "REVESTIMENTO TIPO: " + cabecalho["Tipo Revest."]
    sheet.cell(row = 4, column = 17).value = "DATA: " + cabecalho["Data"]
    sheet.cell(row = 4, column = 26).value = cabecalho["km ini"]
    sheet.cell(row = 4, column = 28).value = cabecalho["km fim"]

    sheet.print_area = 'A1:AC' + str(planilha.shape[0] + start_row - 1)
    book.save(nomeIGG)
    print("    ... pronto para uso \n")


def AnexoANTT(main_path, modelo_path, file, cabecalho, planilha, Valores_Ficha, num_segmentos):
    # Ajustando x para X
    for column in ['OK', 'Fi', 'TTC', 'TTL', 'TLC', 'TLL', 'TRR', 'J', 'TB', 'JE', 'TBE', 'ALP', 'ATP', 'ALC', 'ATC', 'O', 'P', 'E', 'Ex', 'D', 'R', "OK-ATR"]:
        planilha[column].where(planilha[column] != "x", "X", inplace=True)
    for atr in ['ATR Dir (mm)', 'ATR Esq (mm)']:
        planilha[atr].where(planilha[atr] != None, "-", inplace=True)

    # Abrindo arquivo modelo
    new_name = (file.split("PP_")[1]).split(".")[0] + ".xlsx"
    print("Iniciando a exportação do arquivo: ANTT - " + new_name)
    nomeIGG = os.path.join(main_path, 'ANTT - ' + new_name)
    shutil.copy(modelo_path, nomeIGG)
    book = op.load_workbook(nomeIGG)

    # Formatação do padrão ANTT - ANEXO A
    # Editando arquivo
    sheet = book.active
    sheet = book['Anexo A']
    start_col = 1
    start_row = 8
    end_col = 27

    # Copia formatação
    for line in range(planilha.shape[0]):
        for col in range(end_col):
            sheet.cell(row = start_row + line, column = start_col + col)._style = copy(sheet.cell(row = start_row, column = col+1)._style)
            sheet.row_dimensions[start_row + line].height = 20.5
        
    anexo_colunas = ['Início', "OK", "Secao Terra.", 'Fi', 'TTC', 'TTL', 'TLC', 'TLL', 'TRR', 'J', 'TB', 'JE', 'TBE', 'ALP', 'ATP', 'ALC', 'ATC', 'O', 'P', 'E', 'Ex', 'D', 'R', 'ATR Esq (mm)', 'ATR Dir (mm)', "OK-ATR", 'Observação']
    for line in range(planilha.shape[0]):
        for col in range(len(anexo_colunas)):
            sheet.cell(row = start_row + line, column = start_col + col).value = planilha[anexo_colunas[col]][line]
        
    sheet.cell(row = 4, column = 3).value = cabecalho["Rodovia"]
    sheet.cell(row = 5, column = 3).value = cabecalho["Data"]
    sheet.cell(row = 4, column = 9).value = cabecalho["Tipo pista"]
    sheet.cell(row = 5, column = 9).value = cabecalho["km ini"]
    sheet.cell(row = 4, column = 15).value = cabecalho["Faixa"]
    sheet.cell(row = 5, column = 15).value = cabecalho["km fim"]
    sheet.cell(row = 4, column = 21).value = cabecalho["Sentido"]
    sheet.cell(row = 5, column = 21).value = cabecalho["Operador"]

    sheet.print_area = 'A1:AA' + str(planilha.shape[0] + start_row - 1)
    print("    Anexo A - OK")

    # Formatação do padrão ANTT - ANEXO B
    sheet = book.active
    sheet = book['Anexo B']
    start_col = 1
    start_row = 4
    end_col = 9
    end_row = 28

    # Copia formatação
    next_line = start_row
    for segmento in range(num_segmentos - 1):
        for line in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                sheet.cell(row = (next_line + (end_row - start_row) + 1), column = col)._style = copy(sheet.cell(row = line, column = col)._style)
                sheet.cell(row = (next_line + (end_row - start_row) + 1), column = col).value = sheet.cell(row = line, column = col).value
            next_line = next_line + 1
    
    # Copia células mescladas
                      # Coluna, linha, coluna, linha
    lista_mescladas = [(1,9,1,10), (2,9,2,10), (3,9,4,9), (3,10,4,10), (7,9,8,9), (7,10,8,10), (9,9,9,10), (3,11,4,11), 
                       (3,12,4,12), (3,13,4,13), (3,14,4,14), (3,15,4,15), (3,16,4,16), (3,17,4,17), (3,18,4,18), 
                       (5,14,5,18), (7,11,8,11), (7,12,8,12), (7,13,8,13), (7,14,8,14), (7,15,8,15), (7,16,8,16), 
                       (7,17,8,17), (7,18,8,18), (1,19,1,22), (2,19,2,22), (3,20,3,22), (4,20,4,22), (5,19,5,22), 
                       (6,19,6,20), (6,21,6,22), (7,19,7,20), (7,21,7,22), (8,19,8,20), (8,21,8,22), (9,19,9,22), 
                       (1,23,1,26), (2,23,2,26), (3,24,3,26), (4,24,4,26), (5,23,5,26), (6,23,6,24), (6,25,6,26), 
                       (7,23,7,24), (7,25,7,26), (8,23,8,24), (8,25,8,26), (9,23,9,26), (1,27,2,28), (3,27,4,28), 
                       (5,27,8,27), (5,28,8,28)]
    for segmento in range(1, num_segmentos):
        for celula in lista_mescladas:
            min_row = celula[1] + ((end_row - start_row) + 1) * segmento
            max_row = celula[3] + ((end_row - start_row) + 1) * segmento
            sheet.merge_cells(start_row = min_row, start_column = celula[0], end_row = max_row, end_column = celula[2])
    
    # Monta ficha
    next_line = start_row
    for segmento in range(num_segmentos):
        sheet.cell(row = next_line + 1, column = 2).value = cabecalho["Rodovia"]
        sheet.cell(row = next_line + 2, column = 2).value = cabecalho["Trecho"]
        sheet.cell(row = next_line + 3, column = 2).value = cabecalho["Faixa"]
        sheet.cell(row = next_line + 1, column = 9).value = cabecalho["Tipo pista"]
        sheet.cell(row = next_line + 2, column = 9).value = str(Valores_Ficha[segmento]['Início']) + " a " + str(Valores_Ficha[segmento]['Fim'])
        
        for item in range(8):
            sheet.cell(row = next_line + item + 7, column = 3).value = Valores_Ficha[segmento]['Freq. Absoluta'][item]
            sheet.cell(row = next_line + item + 7, column = 6).value = Valores_Ficha[segmento]['Freq. Relativa'][item]
        for item in range(3):
            sheet.cell(row = next_line + item + 7, column = 5).value = Valores_Ficha[segmento]['Freq. Abs. Consi.'][item]
        for item in range(9):
            sheet.cell(row = next_line + item + 7, column = 9).value = Valores_Ficha[segmento]['IGI'][item]
        sheet.cell(row = next_line + 19, column = 9).value = Valores_Ficha[segmento]['IGI'][9]
        sheet.cell(row = next_line + 23, column = 9).value = Valores_Ficha[segmento]['IGG']
        sheet.cell(row = next_line + 24, column = 9).value = Valores_Ficha[segmento]['Conceito']

        # Duplicado para poder puxar valores no resumo
        sheet.cell(row = next_line + 23, column = 12).value = Valores_Ficha[segmento]['IGG']
        sheet.cell(row = next_line + 24, column = 12).value = Valores_Ficha[segmento]['Conceito']

        sheet.cell(row = next_line + 23, column = 3).value = Valores_Ficha[segmento]['Num. Estacoes']
        sheet.cell(row = next_line + 16, column = 4).value = Valores_Ficha[segmento]['Med. Dir.']
        sheet.cell(row = next_line + 16, column = 3).value = Valores_Ficha[segmento]['Med. Esq.']
        sheet.cell(row = next_line + 17, column = 6).value = (Valores_Ficha[segmento]['Med. Esq.'] + Valores_Ficha[segmento]['Med. Dir.'])/2
        sheet.cell(row = next_line + 20, column = 4).value = Valores_Ficha[segmento]['Var. Dir.']
        sheet.cell(row = next_line + 20, column = 3).value = Valores_Ficha[segmento]['Var. Esq.']
        sheet.cell(row = next_line + 21, column = 6).value = (Valores_Ficha[segmento]['Var. Esq.'] + Valores_Ficha[segmento]['Var. Dir.'])/2
        # Ajustado altura das linhas
        for line in range(25):
            sheet.row_dimensions[next_line + line].height = 19.75

        next_line = next_line + (end_row - start_row) + 1
    
    sheet.print_area = 'A1:I' + str(next_line - 1)
    print("    Anexo B - OK")

    # Salvando arquivo editado
    book.save(nomeIGG)
    print("    ... pronto para uso \n")


def AnexoBPavesys(main_path, modelo_path, file, Valores_Ficha, num_segmentos, cabecalho):
    # Abrindo arquivo modelo
    new_name = (file.split("PP_")[1]).split(".")[0] + ".xlsx"
    print("Iniciando a exportação do arquivo: Anexo B - Pavesys - " + new_name)
    nomeIGG = os.path.join(main_path, 'Anexo B - Pavesys - ' + new_name)
    shutil.copy(modelo_path, nomeIGG)
    book = op.load_workbook(nomeIGG)

    # Editando arquivo
    sheet = book.active
    sheet.title = '_'
    sheet = book['_']
    start_col = 1
    start_row = 6
    end_col = 7
    end_row = 27

    # Copia formatação
    next_line = start_row
    for segmento in range(num_segmentos - 1):
        for line in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                sheet.cell(row = (next_line + (end_row - start_row) + 1), column = col)._style = copy(sheet.cell(row = line, column = col)._style)
                sheet.cell(row = (next_line + (end_row - start_row) + 1), column = col).value = sheet.cell(row = line, column = col).value
            next_line = next_line + 1
    # Monta ficha
    next_line = start_row
    for segmento in range(num_segmentos):
        sheet.cell(row = next_line, column = 2).value = segmento + 1
        sheet.cell(row = next_line, column = 4).value = Valores_Ficha[segmento]['Início']
        sheet.cell(row = next_line, column = 6).value = Valores_Ficha[segmento]['Fim']
        for item in range(8):
            sheet.cell(row = next_line + item + 2, column = 3).value = Valores_Ficha[segmento]['Freq. Absoluta'][item]
            sheet.cell(row = next_line + item + 2, column = 5).value = Valores_Ficha[segmento]['Freq. Relativa'][item]
        for item in range(3):
            sheet.cell(row = next_line + item + 2, column = 4).value = Valores_Ficha[segmento]['Freq. Abs. Consi.'][item]
        for item in range(10):
            sheet.cell(row = next_line + item + 2, column = 7).value = Valores_Ficha[segmento]['IGI'][item]
        sheet.cell(row = next_line + 12, column = 7).value = Valores_Ficha[segmento]['IGG']
        sheet.cell(row = next_line + 13, column = 7).value = Valores_Ficha[segmento]['Conceito']
        sheet.cell(row = next_line + 12, column = 3).value = Valores_Ficha[segmento]['Num. Estacoes']
        sheet.cell(row = next_line + 18, column = 4).value = Valores_Ficha[segmento]['Med. Dir.']
        sheet.cell(row = next_line + 18, column = 3).value = Valores_Ficha[segmento]['Med. Esq.']
        sheet.cell(row = next_line + 18, column = 6).value = Valores_Ficha[segmento]['Desv. Dir.']
        sheet.cell(row = next_line + 18, column = 5).value = Valores_Ficha[segmento]['Desv. Esq.']
        sheet.cell(row = next_line + 18, column = 7).value = (Valores_Ficha[segmento]['Desv. Esq.'] + Valores_Ficha[segmento]['Desv. Dir.'])/2
        sheet.cell(row = next_line + 20, column = 3).value = (Valores_Ficha[segmento]['Med. Esq.'] + Valores_Ficha[segmento]['Med. Dir.'])/2
        sheet.cell(row = next_line + 20, column = 6).value = Valores_Ficha[segmento]['Var. Dir.']
        sheet.cell(row = next_line + 20, column = 5).value = Valores_Ficha[segmento]['Var. Esq.']
        sheet.cell(row = next_line + 20, column = 7).value = (Valores_Ficha[segmento]['Var. Esq.'] + Valores_Ficha[segmento]['Var. Dir.'])/2
        next_line = next_line + (end_row - start_row) + 1

        sheet.cell(row = 2, column = 1).value = "RODOVIA: " + cabecalho["Rodovia"]
        sheet.cell(row = 3, column = 1).value = "TRECHO: " + cabecalho["STH"]
        sheet.cell(row = 4, column = 1).value = "PISTA: " + cabecalho["Tipo pista"]
        sheet.cell(row = 5, column = 1).value = "FAIXA: " + cabecalho["Faixa"]
        sheet.cell(row = 5, column = 2).value = "SENTIDO: " + cabecalho["Sentido"]
        sheet.cell(row = 2, column = 3).value = "Data: " + cabecalho["Data"]

    # Salvando arquivo editado
    sheet.print_area = 'A1:G' + str(next_line - 2)
    book.save(nomeIGG)
    print("    ... pronto para uso \n")


def AnexoBDERSP(main_path, modelo_path, file, Valores_Ficha, num_segmentos, cabecalho, modelo_path_unifilar, planilha):
    # Abrindo arquivo modelo
    new_name = (file.split("PP_")[1]).split(".")[0] + ".xlsx"
    print("Iniciando a exportação do arquivo: Anexo B - DER-SP - " + new_name)
    nomeIGG = os.path.join(main_path, 'Anexo B - DER-SP - ' + new_name)
    shutil.copy(modelo_path, nomeIGG)
    book = op.load_workbook(nomeIGG)

    # Editando arquivo
    sheet = book['_']
    start_col = 1
    start_row = 6
    end_col = 7
    end_row = 27
    
    # Copia formatação
    next_line = start_row
    for segmento in range(num_segmentos - 1):
        for line in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                sheet.cell(row = (next_line + (end_row - start_row) + 1), column = col)._style = copy(sheet.cell(row = line, column = col)._style)
                sheet.cell(row = (next_line + (end_row - start_row) + 1), column = col).value = sheet.cell(row = line, column = col).value
            next_line = next_line + 1
    # Monta ficha
    next_line = start_row
    for segmento in range(num_segmentos):
        sheet.cell(row = next_line, column = 2).value = segmento + 1
        sheet.cell(row = next_line, column = 4).value = Valores_Ficha[segmento]['Início']*1000
        sheet.cell(row = next_line, column = 6).value = Valores_Ficha[segmento]['Fim']*1000
        for item in range(8):
            sheet.cell(row = next_line + item + 2, column = 3).value = Valores_Ficha[segmento]['Freq. Absoluta'][item]
            sheet.cell(row = next_line + item + 2, column = 5).value = Valores_Ficha[segmento]['Freq. Relativa'][item]
        for item in range(3):
            sheet.cell(row = next_line + item + 2, column = 4).value = Valores_Ficha[segmento]['Freq. Abs. Consi.'][item]
        for item in range(10):
            sheet.cell(row = next_line + item + 2, column = 7).value = Valores_Ficha[segmento]['IGI'][item]
        sheet.cell(row = next_line + 12, column = 7).value = Valores_Ficha[segmento]['IGG']
        sheet.cell(row = next_line + 13, column = 7).value = Valores_Ficha[segmento]['Conceito']
        sheet.cell(row = next_line + 12, column = 3).value = Valores_Ficha[segmento]['Num. Estacoes']
        sheet.cell(row = next_line + 18, column = 4).value = Valores_Ficha[segmento]['Med. Dir.']
        sheet.cell(row = next_line + 18, column = 3).value = Valores_Ficha[segmento]['Med. Esq.']
        sheet.cell(row = next_line + 18, column = 6).value = Valores_Ficha[segmento]['Desv. Dir.']
        sheet.cell(row = next_line + 18, column = 5).value = Valores_Ficha[segmento]['Desv. Esq.']
        sheet.cell(row = next_line + 18, column = 7).value = (Valores_Ficha[segmento]['Desv. Esq.'] + Valores_Ficha[segmento]['Desv. Dir.'])/2
        sheet.cell(row = next_line + 20, column = 3).value = (Valores_Ficha[segmento]['Med. Esq.'] + Valores_Ficha[segmento]['Med. Dir.'])/2
        sheet.cell(row = next_line + 20, column = 6).value = Valores_Ficha[segmento]['Var. Dir.']
        sheet.cell(row = next_line + 20, column = 5).value = Valores_Ficha[segmento]['Var. Esq.']
        sheet.cell(row = next_line + 20, column = 7).value = (Valores_Ficha[segmento]['Var. Esq.'] + Valores_Ficha[segmento]['Var. Dir.'])/2
        next_line = next_line + (end_row - start_row) + 1

        sheet.cell(row = 2, column = 2).value = cabecalho["Rodovia"]
        sheet.cell(row = 3, column = 2).value = cabecalho["STH"]
        sheet.cell(row = 4, column = 2).value = cabecalho["Faixa"]
        sheet.cell(row = 2, column = 3).value = "Data: " + cabecalho["Data"]
        # Ajustado altura das linhas
        for line in range(22):
            sheet.row_dimensions[next_line + line].height = 18.25

    # Salvando arquivo editado
    sheet.print_area = 'A1:G' + str(next_line - 2)
    book.save(nomeIGG)
    print("    ... pronto para uso \n")

    # Abrindo arquivo modelo
    print("Iniciando a exportação do arquivo: Unifilar - DER-SP - " + file)
    nomeIGG = os.path.join(main_path, 'Unifilar - DER-SP - ' + new_name)
    shutil.copy(modelo_path_unifilar, nomeIGG)
    book = op.load_workbook(nomeIGG)

    # Editando arquivo
    sheet = book['_']
    start_col = 2
    start_row = 20
    end_col = 11

    """image_loader = SheetImageLoader(sheet)
    image = image_loader.get('J2')
    image.show()"""

    # Copia formatação
    for line in range(num_segmentos-1):
        for col in range(end_col):
            sheet.cell(row = start_row + line + 1, column = start_col + col)._style = copy(sheet.cell(row = start_row, column = col+2)._style)
    
    sheet.cell(row = 4, column = 2).value = cabecalho["Rodovia"]
    for line in range(num_segmentos):
        sheet.cell(row = line + start_row, column = 2).value = Valores_Ficha[line]['Início']*1000
        sheet.cell(row = line + start_row, column = 3).value = Valores_Ficha[line]['Fim']*1000
        #sheet.cell(row = line, column = 5).value = planilha['Observação'][line]
        sheet.cell(row = line + start_row, column = 7).value = Valores_Ficha[line]['IGG']
        sheet.cell(row = line + start_row, column = 8).value = Valores_Ficha[line]['Conceito']
        sheet.cell(row = line + start_row, column = 10).value = Valores_Ficha[line]['Início']*1000
        sheet.cell(row = line + start_row, column = 11).value = Valores_Ficha[line]['Fim']*1000

    sheet.print_area = 'A1:L' + str(num_segmentos + start_row - 1)
    book.save(nomeIGG)
    print("    ... pronto para uso \n")


def GeraP21(main_path, file, Valores_Ficha, num_segmentos, crescente):
    # Abrindo arquivo modelo
    new_name = 'P21 - ' + (file.split("PP_")[1]) + ".xlsx"
    print("Iniciando a exportação do arquivo: " + new_name)
    nomeIGG = os.path.join(main_path, new_name)
  
    # Monta ficha
    dict_p21 = {'km ini': [], 'km fim': [], 'ATR': [], 'IGG': [], 'FC-2': [], 'FC-3': [], "km referencia": [], "Nome ramo": [], "Faixa": []}
    for segmento in range(num_segmentos):
        dict_p21['km ini'].append(Valores_Ficha[segmento]['Início'])
        dict_p21['km fim'].append(Valores_Ficha[segmento]['Fim'])
        dict_p21['ATR'].append((Valores_Ficha[segmento]['Med. Esq.'] + Valores_Ficha[segmento]['Med. Dir.'])/2)
        dict_p21['IGG'].append(Valores_Ficha[segmento]['IGG'])
        dict_p21['FC-2'].append(Valores_Ficha[segmento]['FC-2'])
        dict_p21['FC-3'].append(Valores_Ficha[segmento]['FC-3'])
        dict_p21["km referencia"].append(Valores_Ficha[segmento]["km referencia"])
        dict_p21["Nome ramo"].append(Valores_Ficha[segmento]["Nome ramo"])
        dict_p21["Faixa"].append(Valores_Ficha[segmento]["Faixa"])

    df_export = pd.DataFrame.from_dict(dict_p21)
    df_export.to_excel(os.path.join(main_path, new_name), index=False)

    print("    ... pronto para uso \n")


def SeparaDF (df):
    # Criar uma lista de dataframes separados
    dfs_separados = []
    df_atual = pd.DataFrame()

    # Iterar sobre as linhas do dataframe original
    for index, row in df.iterrows():
        # Verificar se a linha está em branco
        if row.isnull().all():
            # Adicionar o dataframe atual à lista de dataframes separados
            if not df_atual.empty:
                dfs_separados.append(df_atual)
                df_atual = pd.DataFrame()
        else:
            # Adicionar a linha atual ao dataframe atual
            df_atual = pd.concat([df_atual, df.loc[[index]]])

    # Adicionar o último dataframe atual à lista de dataframes separados
    if not df_atual.empty:
        dfs_separados.append(df_atual)
    return dfs_separados


# ------------------------ FUNÇÕES PARA FOR ------------------------
def ForDecrescente(path, file_loc, df_original, df_original_atr, espacamento, artesp, atr_name, pista):
    km_decrescente = df_original['Início'].tolist()
    estacoes_decrescente = CriaEstacoes(km_decrescente, espacamento)
    trechos_trincamento_decrescente = CriaEstacoes(km_decrescente, 0.01)
    df_LVCdecrescente, teclas_var = PadraoLVC(df_original)
    if teclas_var and artesp:
        df_area_decrescente = AreaTrincada(file_loc, trechos_trincamento_decrescente, max(km_decrescente), min(km_decrescente))
        df_decrescente_filtrado, df_decrescente_absoluto = FiltroEstacoes(estacoes_decrescente, df_LVCdecrescente, df_original_atr, df_area_decrescente, False, artesp, teclas_var)
    else:
        df_area_decrescente = pd.DataFrame()
        df_decrescente_filtrado, df_decrescente_absoluto = FiltroEstacoes(estacoes_decrescente, df_LVCdecrescente, df_original_atr, df_area_decrescente, False, artesp, teclas_var)
    df_decrescente_filtrado.sort_values(by=['Início'], ascending=False, inplace=True)
    df_decrescente_absoluto.sort_values(by=['Início'], ascending=False, inplace=True)
    cabecalho_decrescente = DadosCabecalho(path, pista, atr_name, df_original)
    df_decrescente_filtrado = ColunasFaltantes(df_decrescente_filtrado, cabecalho_decrescente)
    df_decrescente_absoluto = ColunasFaltantes(df_decrescente_absoluto, cabecalho_decrescente)
    return df_decrescente_filtrado, df_decrescente_absoluto, cabecalho_decrescente


def ForCrescente(path, file_loc, df_original, df_original_atr, espacamento, artesp, atr_name, pista):
    km_crescente = df_original['Início'].tolist()
    estacoes_crescente = CriaEstacoes(km_crescente, espacamento)
    trechos_trincamento_crescente = CriaEstacoes(km_crescente, 0.01)
    df_LVCcrescente, teclas_var = PadraoLVC(df_original)
    if teclas_var and artesp:
        df_area_crescente = AreaTrincada(file_loc, trechos_trincamento_crescente, min(km_crescente), max(km_crescente))
        df_crescente_filtrado, df_crescente_absoluto = FiltroEstacoes(estacoes_crescente, df_LVCcrescente, df_original_atr, df_area_crescente, True, artesp, teclas_var)
    else:
        df_area_crescente = pd.DataFrame()
        df_crescente_filtrado, df_crescente_absoluto = FiltroEstacoes(estacoes_crescente, df_LVCcrescente, df_original_atr, df_area_crescente, True, artesp, teclas_var)
    df_crescente_filtrado.sort_values('Início', ascending=True, inplace=True)
    df_crescente_absoluto.sort_values('Início', ascending=True, inplace=True)
    cabecalho_crescente = DadosCabecalho(path, pista, atr_name, df_original)
    df_crescente_filtrado = ColunasFaltantes(df_crescente_filtrado, cabecalho_crescente)
    df_crescente_absoluto = ColunasFaltantes(df_crescente_absoluto, cabecalho_crescente)

    # Zerar km se estiver processando um ramo
    if "RAMO" in pista:
        max_filtrado = df_crescente_filtrado['Início'].min()
        max_absoluto = df_crescente_absoluto['Início'].min()
        df_crescente_filtrado['Início'] = df_crescente_filtrado['Início'] - max_filtrado
        df_crescente_absoluto['Início'] = df_crescente_absoluto['Início'] - max_absoluto
        df_crescente_filtrado['Fim'] = df_crescente_filtrado['Fim'] - max_filtrado
        df_crescente_absoluto['Fim'] = df_crescente_absoluto['Fim'] - max_absoluto

    return df_crescente_filtrado, df_crescente_absoluto, cabecalho_crescente


# ------------------------ CÓDIGO FUNÇÃO PRINCIPAL ------------------------
def IGG(path, antt, artesp, dersp, pavesys, pavesys_faixa, segmento_homogeneo):

    # Modelo formatação IGG
    # Modelos Antt
    modelo_path_dupla = r'C:\Pavesys\Templates\Padrão ANTT\IGG Pista Dupla.xlsx'
    modelo_path_simples = r'C:\Pavesys\Templates\Padrão ANTT\IGG Pista Simples.xlsx'
    # Modelo Artesp
    modelo_path_p15 = r'C:\Pavesys\Templates\Padrão ARTESP\P15 - IGG.xlsx'
    # Modelos Der-SP
    modelo_path_anexoA_SP = r'C:\Pavesys\Templates\Padrão DER SP\Planilha IGG SP.xlsx'
    modelo_path_anexoB_SP = r'C:\Pavesys\Templates\Padrão DER SP\Cálculo IGG SP.xlsx'
    modelo_path_unifilar = r'C:\Pavesys\Templates\Padrão DER SP\Planilha Unifilar SP.xlsx'
    # Modelos Pavesys
    modelo_path_anexoA = r'C:\Pavesys\Templates\Padrão PAVESYS\Planilha IGG.xlsx'
    modelo_path_anexoB = r'C:\Pavesys\Templates\Padrão PAVESYS\Ficha IGG.xlsx'

    # Organizando arquivos para processamento
    dc_pista_simples, dc_pista_dupla_cresc, dc_pista_dupla_decresc, dc_pista_adc_cres, dc_pista_adc_decres, dc_pista_ramo, atr_pista_simples, atr_pista_dupla_cresc, atr_pista_dupla_decresc, atr_pista_adc_cres, atr_pista_adc_decres, atr_pista_ramo = leitor.dicionario_arquivos(path)
    # Df que armazenara todos os dados pra compor a P15
    df_p15 = pd.DataFrame()

    # PISTA SIMPLES - ATENÇÃO: PRESUME-SE QUE O DF DA CRESCENTE E DECRESCENTE TENHAM O MESMO TAMANHO
    if bool(dc_pista_simples):

        for chave in dc_pista_simples:
            for pista in dc_pista_simples[chave]:
                for arq in atr_pista_simples[chave]:
                    if arq.split("_ATR.")[0] == pista.split(".")[0]:
                        atr_name = arq
                file_loc, df_original, df_original_atr = ExtrairDF(path, pista, atr_name)
                # Invertendo pistas decrescentes, mudando para padrão LVC e criando estações para cada pista
                if pista.split("_")[3] == "D":
                    df_decrescente_filtrado, df_decrescente_absoluto, cabecalho_decrescente = ForDecrescente(path, file_loc, df_original, df_original_atr, 0.02, artesp, atr_name, pista)
                    km_decrescente = df_original['Início'].tolist()
                    km_decrescente_fim = df_original['Fim'].tolist()
                if pista.split("_")[3] == "C":
                    df_crescente_filtrado, df_crescente_absoluto, cabecalho_crescente = ForCrescente(path, file_loc, df_original, df_original_atr, 0.02, artesp, atr_name, pista)
                    km_crescente = df_original['Início'].tolist()
                    km_crescente_fim = df_original['Fim'].tolist()
                    estacoes_crescente = CriaEstacoes(km_crescente, 0.02)
            
            # unir pistas e intercalando
            df_concatenado, df_concat_crescente, df_concat_decrescente = UnirPistaSimples(df_crescente_filtrado, df_decrescente_filtrado, estacoes_crescente)
            df_absoluto_concatenado, df_abs_concat_crescente, df_abs_concat_decrescente = UnirPistaSimples(df_crescente_absoluto, df_decrescente_absoluto, estacoes_crescente)

            # Exporta inventário no modelo pavesys e dersp
            if pavesys or pavesys_faixa:
                AnexoAPavesys(path, modelo_path_anexoA, 'PP_'+chave+'_S_C_1.xlsx', cabecalho_crescente, df_abs_concat_crescente)
                
                # APAGAR DEPOIS
                #for column in ['OK', 'Fi', 'TTC', 'TTL', 'TLC', 'TLL', 'TRR', 'J', 'TB', 'JE', 'TBE', 'ALP', 'ATP', 'O', 'P', 'Ex', 'D', 'R', 'ALC', 'ATC', 'E']:
                #    df_abs_concat_crescente[column].where(df_abs_concat_crescente[column] != "X", "x", inplace=True)
                #df_abs_concat_crescente.to_excel(os.path.join(path, "df_abs_concat_crescente.xlsx"))

                AnexoAPavesys(path, modelo_path_anexoA, 'PP_'+chave+'_S_D_1.xlsx', cabecalho_decrescente, df_abs_concat_decrescente)
            if dersp:
                AnexoADERSP(path, modelo_path_anexoA_SP, 'PP_'+chave+'_S_C_1.xlsx', cabecalho_crescente, df_abs_concat_crescente)
                AnexoADERSP(path, modelo_path_anexoA_SP, 'PP_'+chave+'_S_D_1.xlsx', cabecalho_decrescente, df_abs_concat_decrescente)

            # Cria seções para cálculo de IGG
            secoes_km = CriaEstacoes(km_crescente, segmento_homogeneo)
            if len(secoes_km) == 0:
                secoes_km.append(min(km_crescente))
            if secoes_km[0] != km_crescente[0]:
                secoes_km.insert(0, round(km_crescente[0], 3))
            if secoes_km[-1] != round(km_crescente_fim[-1], 3):
                secoes_km.append(round(km_crescente_fim[-1], 3))
            
            secoes_km_decres = CriaEstacoes(km_decrescente, segmento_homogeneo)
            if len(secoes_km_decres) == 0:
                secoes_km_decres.append(max(km_decrescente))
            if secoes_km_decres[0] != km_decrescente[0]:
                secoes_km_decres.insert(0, km_decrescente[0])
            if secoes_km_decres[-1] != round(km_decrescente_fim[-1], 3):
                secoes_km_decres.append(round(km_decrescente_fim[-1], 3))

            # Exporta fixa no modelo pavesys e dersp
            if pavesys or dersp:
                Valores_Ficha = {}
                for secao in range(len(secoes_km) - 1):
                    df = df_concatenado[(df_concatenado['Início'] >= secoes_km[secao]) & (df_concatenado['Início'] < secoes_km[secao + 1])]
                    num_estacoes = df['Início'].count()
                    df_abs_concat = df_absoluto_concatenado[(df_absoluto_concatenado['Início'] >= secoes_km[secao]) & (df_absoluto_concatenado['Início'] < secoes_km[secao + 1])]
                    Valores_Ficha[secao] = FichaIGG(df, df_abs_concat, num_estacoes, [secoes_km[secao], secoes_km[secao + 1]])
                if pavesys:
                    AnexoBPavesys(path, modelo_path_anexoB, 'PP_'+chave+'.xlsx', Valores_Ficha, len(secoes_km) - 1, cabecalho_crescente)
                if dersp:
                    AnexoBDERSP(path, modelo_path_anexoB_SP, 'PP_'+chave+'.xlsx', Valores_Ficha, len(secoes_km) - 1, cabecalho_crescente, modelo_path_unifilar, df_abs_concat_crescente)
            
            if antt or artesp or pavesys_faixa:
                # Formato ANTT para faixa crescente
                Valores_Ficha = {}
                for secao in range(len(secoes_km) - 1):
                    df = df_concat_crescente[(df_concat_crescente['Início'] >= secoes_km[secao]) & (df_concat_crescente['Início'] < secoes_km[secao + 1])]
                    num_estacoes = df['Início'].count()
                    df_abs_concat = df_abs_concat_crescente[(df_abs_concat_crescente['Início'] >= secoes_km[secao]) & (df_abs_concat_crescente['Início'] < secoes_km[secao + 1])]
                    Valores_Ficha[secao] = FichaIGG(df, df_abs_concat, num_estacoes, [secoes_km[secao], secoes_km[secao + 1]])
                if antt:
                    AnexoANTT(path, modelo_path_simples, 'PP_'+chave+'_S_C_1.xlsx', cabecalho_crescente, df_abs_concat_crescente, Valores_Ficha, len(secoes_km) - 1)
                if pavesys_faixa:
                    AnexoBPavesys(path, modelo_path_anexoB, 'PP_'+chave+'_S_C_1.xlsx', Valores_Ficha, len(secoes_km) - 1, cabecalho_crescente)
                if artesp:
                    GeraP21(path, 'PP_'+chave+'_S_C_1', Valores_Ficha, len(secoes_km) - 1, True)
                
                # Formato ANTT para faixa decrescente
                Valores_Ficha = {}
                for secao in range(len(secoes_km_decres) - 1):
                    df = df_concat_decrescente[(df_concat_decrescente['Início'] <= secoes_km_decres[secao]) & (df_concat_decrescente['Início'] > secoes_km_decres[secao + 1])]
                    num_estacoes = df['Início'].count()
                    df_abs_concat = df_abs_concat_decrescente[(df_abs_concat_decrescente['Início'] <= secoes_km_decres[secao]) & (df_abs_concat_decrescente['Início'] > secoes_km_decres[secao + 1])]
                    Valores_Ficha[secao] = FichaIGG(df, df_abs_concat, num_estacoes, [secoes_km_decres[secao], secoes_km_decres[secao + 1]])

                df_abs_concat_decrescente.sort_values(by=['Início'], ascending=False, inplace=True, ignore_index=True)
                if antt:
                    AnexoANTT(path, modelo_path_simples, 'PP_'+chave+'_S_D_1.xlsx', cabecalho_decrescente, df_abs_concat_decrescente, Valores_Ficha, len(secoes_km_decres) - 1)
                if pavesys_faixa:
                    AnexoBPavesys(path, modelo_path_anexoB, 'PP_'+chave+'_S_D_1.xlsx', Valores_Ficha, len(secoes_km_decres) - 1, cabecalho_decrescente)
                if artesp:
                    GeraP21(path, 'PP_'+chave+'_S_D_1', Valores_Ficha, len(secoes_km_decres) - 1, False)
            
            # Armazenando valores para P15
            if artesp:
                df_p15 = pd.concat([df_p15, df_abs_concat_crescente, df_abs_concat_decrescente], ignore_index = True)
                df_p15.to_excel(os.path.join(path, "copia de seguranca - P15.xlsx"))
                

    # PISTA DUPLA CRESCENTE ------------------------------------------------------------------------------
    if bool(dc_pista_dupla_cresc):

        for chave in dc_pista_dupla_cresc:
            for pista in dc_pista_dupla_cresc[chave]:
                for arq in atr_pista_dupla_cresc[chave]:
                    if arq.split("_ATR.")[0] == pista.split(".")[0]:
                        atr_name = arq
                file_loc, df_original, df_original_atr = ExtrairDF(path, pista, atr_name)
                # Invertendo pistas decrescentes, mudando para padrão LVC e criando estações para cada pista
                if int((pista.split(".")[0]).split("_")[4]) == 1:
                    df_crescente_filtrado, df_crescente_absoluto, cabecalho_crescente = ForCrescente(path, file_loc, df_original, df_original_atr, 0.04, artesp, atr_name, pista)
                    km_crescente = df_original['Início'].tolist()
                    km_crescente_fim = df_original['Fim'].tolist()
                    estacoes_crescente = CriaEstacoes(km_crescente, 0.04)
                else:
                    df_crescente_filtrado, df_crescente_absoluto, cabecalho_crescente = ForCrescente(path, file_loc, df_original, df_original_atr, 0.02, artesp, atr_name, pista)
                    km_crescente = df_original['Início'].tolist()
                    km_crescente_fim = df_original['Fim'].tolist()
                    estacoes_crescente = CriaEstacoes(km_crescente, 0.02)
            
                # Exporta inventário no modelo pavesys e dersp
                if pavesys or pavesys_faixa:
                    AnexoAPavesys(path, modelo_path_anexoA, pista.split(".")[0] + '.xlsx', cabecalho_crescente, df_crescente_absoluto)
                if dersp:
                    AnexoADERSP(path, modelo_path_anexoA_SP, pista.split(".")[0] + '.xlsx', cabecalho_crescente, df_crescente_absoluto)

                # Cria seções para cálculo de IGG
                secoes_km = CriaEstacoes(km_crescente, segmento_homogeneo)
                if len(secoes_km) == 0:
                    secoes_km.append(min(km_crescente))
                if secoes_km[0] != km_crescente[0]:
                    secoes_km.insert(0, km_crescente[0])
                if secoes_km[-1] != round(km_crescente_fim[-1], 3):
                    secoes_km.append(round(km_crescente_fim[-1], 3))

                # Exporta fixa no modelo pavesys e dersp
                Valores_Ficha = {}
                for secao in range(len(secoes_km) - 1):
                    df = df_crescente_filtrado[(df_crescente_filtrado['Início'] >= secoes_km[secao]) & (df_crescente_filtrado['Início'] < secoes_km[secao + 1])]
                    num_estacoes = df['Início'].count()
                    df_abs = df_crescente_absoluto[(df_crescente_absoluto['Início'] >= secoes_km[secao]) & (df_crescente_absoluto['Início'] < secoes_km[secao + 1])]
                    Valores_Ficha[secao] = FichaIGG(df, df_abs, num_estacoes, [secoes_km[secao], secoes_km[secao + 1]])
                if pavesys or pavesys_faixa:
                    AnexoBPavesys(path, modelo_path_anexoB, pista.split(".")[0] + '.xlsx', Valores_Ficha, len(secoes_km) - 1, cabecalho_crescente)
                if dersp:
                    AnexoBDERSP(path, modelo_path_anexoB_SP, pista.split(".")[0] + '.xlsx', Valores_Ficha, len(secoes_km) - 1, cabecalho_crescente, modelo_path_unifilar, df_crescente_absoluto)
                if antt:
                    AnexoANTT(path, modelo_path_simples, pista.split(".")[0] + '.xlsx', cabecalho_crescente, df_crescente_absoluto, Valores_Ficha, len(secoes_km) - 1)
                
                # Armazenando valores para P15
                if artesp:
                    df_p15 = pd.concat([df_p15, df_crescente_absoluto], ignore_index = True)
                    df_p15.to_excel(os.path.join(path, "copia de seguranca - P15.xlsx"))
                    GeraP21(path, pista.split(".")[0], Valores_Ficha, len(secoes_km) - 1, True)


    # PISTA DUPLA DECRESCENTE ----------------------------------------------------------------------------
    if bool(dc_pista_dupla_decresc):

        for chave in dc_pista_dupla_decresc:
            for pista in dc_pista_dupla_decresc[chave]:
                for arq in atr_pista_dupla_decresc[chave]:
                    if arq.split("_ATR.")[0] == pista.split(".")[0]:
                        atr_name = arq
                file_loc, df_original, df_original_atr = ExtrairDF(path, pista, atr_name)
                # Invertendo pistas decrescentes, mudando para padrão LVC e criando estações para cada pista
                if int((pista.split(".")[0]).split("_")[4]) == 1:
                    df_decrescente_filtrado, df_decrescente_absoluto, cabecalho_decrescente = ForDecrescente(path, file_loc, df_original, df_original_atr, 0.04, artesp, atr_name, pista)
                    km_decrescente = df_original['Início'].tolist()
                    km_decrescente_fim = df_original['Fim'].tolist()
                    estacoes_decrescente = CriaEstacoes(km_decrescente, 0.04)
                else:
                    df_decrescente_filtrado, df_decrescente_absoluto, cabecalho_decrescente = ForDecrescente(path, file_loc, df_original, df_original_atr, 0.02, artesp, atr_name, pista)
                    km_decrescente = df_original['Início'].tolist()
                    km_decrescente_fim = df_original['Fim'].tolist()
                    estacoes_decrescente = CriaEstacoes(km_decrescente, 0.02)
            
                # Exporta inventário no modelo pavesys e dersp
                if pavesys or pavesys_faixa:
                    AnexoAPavesys(path, modelo_path_anexoA, pista.split(".")[0] + '.xlsx', cabecalho_decrescente, df_decrescente_absoluto)
                if dersp:
                    AnexoADERSP(path, modelo_path_anexoA_SP, pista.split(".")[0] + '.xlsx', cabecalho_decrescente, df_decrescente_absoluto)

                # Cria seções para cálculo de IGG
                secoes_km = CriaEstacoes(km_decrescente, segmento_homogeneo)
                if len(secoes_km) == 0:
                    secoes_km.append(max(km_decrescente))
                if secoes_km[0] != km_decrescente[0]:
                    secoes_km.insert(0, km_decrescente[0])
                if secoes_km[-1] != round(km_decrescente_fim[-1], 3):
                    secoes_km.append(round(km_decrescente_fim[-1], 3))

                # Exporta fixa no modelo pavesys e dersp
                Valores_Ficha = {}
                for secao in range(len(secoes_km) - 1):
                    df = df_decrescente_filtrado[(df_decrescente_filtrado['Início'] <= secoes_km[secao]) & (df_decrescente_filtrado['Início'] > secoes_km[secao + 1])]
                    num_estacoes = df['Início'].count()
                    df_abs = df_decrescente_absoluto[(df_decrescente_absoluto['Início'] <= secoes_km[secao]) & (df_decrescente_absoluto['Início'] > secoes_km[secao + 1])]
                    Valores_Ficha[secao] = FichaIGG(df, df_abs, num_estacoes, [secoes_km[secao], secoes_km[secao + 1]])
                if pavesys or pavesys_faixa:
                    AnexoBPavesys(path, modelo_path_anexoB, pista.split(".")[0] + '.xlsx', Valores_Ficha, len(secoes_km) - 1, cabecalho_decrescente)
                if dersp:
                    AnexoBDERSP(path, modelo_path_anexoB_SP, pista.split(".")[0] + '.xlsx', Valores_Ficha, len(secoes_km) - 1, cabecalho_decrescente, modelo_path_unifilar, df_decrescente_absoluto)
                if antt:
                    AnexoANTT(path, modelo_path_simples, pista.split(".")[0] + '.xlsx', cabecalho_decrescente, df_decrescente_absoluto, Valores_Ficha, len(secoes_km) - 1)
                
                # Armazenando valores para P15
                if artesp:
                    df_p15 = pd.concat([df_p15, df_decrescente_absoluto], ignore_index = True)
                    df_p15.to_excel(os.path.join(path, "copia de seguranca - P15.xlsx"))
                    GeraP21(path, pista.split(".")[0], Valores_Ficha, len(secoes_km) - 1, False)


    # FAIXA ADICIONAL CRESCENTE --------------------------------------------------------------------------
    if bool(dc_pista_adc_cres):

        for chave in dc_pista_adc_cres:
            for pista in dc_pista_adc_cres[chave]:
                for arq in atr_pista_adc_cres[chave]:
                    if arq.split("_ATR.")[0] == pista.split(".")[0]:
                        atr_name = arq
                file_loc, df_original, df_original_atr = ExtrairDF(path, pista, atr_name)
                # Separando dfs pela linha em branco
                lista_df = SeparaDF(df_original)
                # Iniciando processamento
                concat_crescente_filtrado = pd.DataFrame()
                lista_crescente_filtrado = []
                concat_crescente_absoluto = pd.DataFrame()
                lista_crescente_absoluto = []
                lista_cabecalho_crescente = []
                lista_km_crescente = []
                lista_estacoes_crescente = []
                lista_secoes_km = []
                for df_adc in lista_df:
                    df_adc = pd.concat([df_adc], ignore_index=True)
                    df_crescente_filtrado, df_crescente_absoluto, cabecalho_crescente = ForCrescente(path, file_loc, df_adc, df_original_atr, 0.02, artesp, atr_name, pista)
                    km_crescente = df_adc['Início'].tolist()
                    km_crescente_fim = df_adc['Fim'].tolist()
                    estacoes_crescente = CriaEstacoes(km_crescente, 0.02)
                    # Cria seções para cálculo de IGG
                    secoes_km = CriaEstacoes(km_crescente, segmento_homogeneo)
                    if len(secoes_km) == 0:
                        secoes_km.append(min(km_crescente))
                    if secoes_km[0] != min(km_crescente):
                        secoes_km.insert(0, min(km_crescente))
                    if secoes_km[-1] != round(max(km_crescente_fim) , 3):
                        secoes_km.append(round(max(km_crescente_fim), 3))
                    # Armazenando processamento separadamente para cadas adicional
                    concat_crescente_filtrado = pd.concat([concat_crescente_filtrado, df_crescente_filtrado], ignore_index = True)
                    lista_crescente_filtrado.append(df_crescente_filtrado)
                    concat_crescente_absoluto = pd.concat([concat_crescente_absoluto, df_crescente_absoluto], ignore_index = True)
                    lista_crescente_absoluto.append(df_crescente_absoluto)
                    lista_cabecalho_crescente.append(cabecalho_crescente)
                    lista_km_crescente.append(km_crescente)
                    lista_estacoes_crescente.append(estacoes_crescente)
                    lista_secoes_km.append(secoes_km)
                # Exporta inventário no modelo pavesys e dersp
                if pavesys or pavesys_faixa:
                    AnexoAPavesys(path, modelo_path_anexoA, pista.split(".")[0] + '.xlsx', lista_cabecalho_crescente[0], concat_crescente_absoluto)
                if dersp:
                    AnexoADERSP(path, modelo_path_anexoA_SP, pista.split(".")[0] + '.xlsx', lista_cabecalho_crescente[0], concat_crescente_absoluto)
            
                # Exporta fixa no modelo pavesys e dersp
                Valores_Ficha = {}
                contador_ficha = 0
                for adc in range(len(lista_secoes_km)):
                    for secao in range(len(lista_secoes_km[adc]) - 1):
                        df = lista_crescente_filtrado[adc][(lista_crescente_filtrado[adc]['Início'] >= lista_secoes_km[adc][secao]) & (lista_crescente_filtrado[adc]['Início'] < lista_secoes_km[adc][secao + 1])]
                        num_estacoes = df['Início'].count()
                        df_abs = lista_crescente_absoluto[adc][(lista_crescente_absoluto[adc]['Início'] >= lista_secoes_km[adc][secao]) & (lista_crescente_absoluto[adc]['Início'] < lista_secoes_km[adc][secao + 1])]
                        Valores_Ficha[contador_ficha] = FichaIGG(df, df_abs, num_estacoes, [lista_secoes_km[adc][secao], lista_secoes_km[adc][secao + 1]])
                        contador_ficha = contador_ficha + 1
                if pavesys or pavesys_faixa:
                    AnexoBPavesys(path, modelo_path_anexoB, pista.split(".")[0] + '.xlsx', Valores_Ficha, contador_ficha, cabecalho_crescente)
                if dersp:
                    AnexoBDERSP(path, modelo_path_anexoB_SP, pista.split(".")[0] + '.xlsx', Valores_Ficha, contador_ficha, cabecalho_crescente, modelo_path_unifilar, concat_crescente_absoluto)
                if antt:
                    AnexoANTT(path, modelo_path_simples, pista.split(".")[0] + '.xlsx', cabecalho_crescente, concat_crescente_absoluto, Valores_Ficha, contador_ficha)
                    
                # Armazenando valores para P15
                if artesp:
                    df_p15 = pd.concat([df_p15, concat_crescente_absoluto], ignore_index = True)
                    df_p15.to_excel(os.path.join(path, "copia de seguranca - P15.xlsx"))
                    GeraP21(path, pista.split(".")[0], Valores_Ficha, contador_ficha, True)


    # FAIXA ADICIONAL DECRESCENTE ------------------------------------------------------------------------
    if bool(dc_pista_adc_decres):

        for chave in dc_pista_adc_decres:
            for pista in dc_pista_adc_decres[chave]:
                for arq in atr_pista_adc_decres[chave]:
                    if arq.split("_ATR.")[0] == pista.split(".")[0]:
                        atr_name = arq
                file_loc, df_original, df_original_atr = ExtrairDF(path, pista, atr_name)
                # Separando dfs pela linha em branco
                lista_df = SeparaDF(df_original)
                # Iniciando processamento
                concat_decrescente_filtrado = pd.DataFrame()
                lista_decrescente_filtrado = []
                concat_decrescente_absoluto = pd.DataFrame()
                lista_decrescente_absoluto = []
                lista_cabecalho_decrescente = []
                lista_km_decrescente = []
                lista_estacoes_decrescente = []
                lista_secoes_km = []
                for df_adc in lista_df:
                    df_adc = pd.concat([df_adc], ignore_index=True)
                    df_decrescente_filtrado, df_decrescente_absoluto, cabecalho_decrescente = ForDecrescente(path, file_loc, df_adc, df_original_atr, 0.02, artesp, atr_name, pista)
                    km_decrescente = df_adc['Início'].tolist()
                    km_decrescente_fim = df_adc['Fim'].tolist()
                    estacoes_decrescente = CriaEstacoes(km_decrescente, 0.02)
                    # Cria seções para cálculo de IGG
                    secoes_km = CriaEstacoes(km_decrescente, segmento_homogeneo)
                    if len(secoes_km) == 0:
                        secoes_km.append(max(km_decrescente))
                    if secoes_km[0] != max(km_decrescente):
                        secoes_km.insert(0, max(km_decrescente))
                    if secoes_km[-1] != round(min(km_decrescente_fim), 3):
                        secoes_km.append(round(min(km_decrescente_fim), 3))
                    # Armazenando processamento separadamente para cadas adicional
                    concat_decrescente_filtrado = pd.concat([concat_decrescente_filtrado, df_decrescente_filtrado], ignore_index = True)
                    lista_decrescente_filtrado.append(df_decrescente_filtrado)
                    concat_decrescente_absoluto = pd.concat([concat_decrescente_absoluto, df_decrescente_absoluto], ignore_index = True)
                    lista_decrescente_absoluto.append(df_decrescente_absoluto)
                    lista_cabecalho_decrescente.append(cabecalho_decrescente)
                    lista_km_decrescente.append(km_decrescente)
                    lista_estacoes_decrescente.append(estacoes_decrescente)
                    lista_secoes_km.append(secoes_km)
                # Exporta inventário no modelo pavesys e dersp
                if pavesys or pavesys_faixa:
                    AnexoAPavesys(path, modelo_path_anexoA, pista.split(".")[0] + '.xlsx', lista_cabecalho_decrescente[0], concat_decrescente_absoluto)
                if dersp:
                    AnexoADERSP(path, modelo_path_anexoA_SP, pista.split(".")[0] + '.xlsx', lista_cabecalho_decrescente[0], concat_decrescente_absoluto)
            
                # Exporta fixa no modelo pavesys e dersp
                Valores_Ficha = {}
                contador_ficha = 0
                for adc in range(len(lista_secoes_km)):
                    for secao in range(len(lista_secoes_km[adc]) - 1):
                        df = lista_decrescente_filtrado[adc][(lista_decrescente_filtrado[adc]['Início'] <= lista_secoes_km[adc][secao]) & (lista_decrescente_filtrado[adc]['Início'] > lista_secoes_km[adc][secao + 1])]
                        num_estacoes = df['Início'].count()
                        df_abs = lista_decrescente_absoluto[adc][(lista_decrescente_absoluto[adc]['Início'] <= lista_secoes_km[adc][secao]) & (lista_decrescente_absoluto[adc]['Início'] > lista_secoes_km[adc][secao + 1])]
                        Valores_Ficha[contador_ficha] = FichaIGG(df, df_abs, num_estacoes, [lista_secoes_km[adc][secao], lista_secoes_km[adc][secao + 1]])
                        contador_ficha = contador_ficha + 1
                if pavesys or pavesys_faixa:
                    AnexoBPavesys(path, modelo_path_anexoB, pista.split(".")[0] + '.xlsx', Valores_Ficha, contador_ficha, cabecalho_decrescente)
                if dersp:
                    AnexoBDERSP(path, modelo_path_anexoB_SP, pista.split(".")[0] + '.xlsx', Valores_Ficha, contador_ficha, cabecalho_decrescente, modelo_path_unifilar, concat_decrescente_absoluto)
                if antt:
                    AnexoANTT(path, modelo_path_simples, pista.split(".")[0] + '.xlsx', cabecalho_decrescente, concat_decrescente_absoluto, Valores_Ficha, contador_ficha)
                    
                # Armazenando valores para P15
                if artesp:
                    df_p15 = pd.concat([df_p15, concat_decrescente_absoluto], ignore_index = True)
                    df_p15.to_excel(os.path.join(path, "copia de seguranca - P15.xlsx"))
                    GeraP21(path, pista.split(".")[0], Valores_Ficha, contador_ficha, False)


    # FAIXA RAMO -----------------------------------------------------------------------------------------
    if bool(dc_pista_ramo):

        for chave in dc_pista_ramo:
            for pista in dc_pista_ramo[chave]:
                for arq in atr_pista_ramo[chave]:
                    if arq.split("_ATR.")[0] == pista.split(".")[0]:
                        atr_name = arq
                file_loc, df_original, df_original_atr = ExtrairDF(path, pista, atr_name)
                # Separando dfs pela linha em branco
                lista_df = SeparaDF(df_original)
                # Iniciando processamento
                concat_crescente_filtrado = pd.DataFrame()
                lista_crescente_filtrado = []
                concat_crescente_absoluto = pd.DataFrame()
                lista_crescente_absoluto = []
                lista_cabecalho_crescente = []
                lista_km_crescente = []
                lista_estacoes_crescente = []
                lista_secoes_km = []
                for df_adc in lista_df:
                    df_adc = pd.concat([df_adc], ignore_index=True)

                    # -----------------------------------
                    # PONTO PARA CORTAR O DF_ORIGINAL_ATR PARA TER APENAS A REFERENTE AO RAMO EM ANÁLISE
                    # -----------------------------------

                    df_crescente_filtrado, df_crescente_absoluto, cabecalho_crescente = ForCrescente(path, file_loc, df_adc, df_original_atr, 0.02, artesp, atr_name, pista)
                    
                    # Zerar km dos ramos
                    min_adc = df_adc['Início'].min()
                    df_adc['Início'] = df_adc['Início'] - min_adc
                    df_adc['Fim'] = df_adc['Fim'] - min_adc

                    km_crescente = df_adc['Início'].tolist()
                    km_crescente_fim = df_adc['Fim'].tolist()
                    estacoes_crescente = CriaEstacoes(km_crescente, 0.02)
                    # Cria seções para cálculo de IGG
                    secoes_km = CriaEstacoes(km_crescente, segmento_homogeneo)
                    if len(secoes_km) == 0:
                        secoes_km.append(min(km_crescente))
                    if secoes_km[0] != min(km_crescente):
                        secoes_km.insert(0, min(km_crescente))
                    if secoes_km[-1] != round(max(km_crescente_fim), 3):
                        secoes_km.append(round(max(km_crescente_fim), 3))
                    # Armazenando processamento separadamente para cadas adicional
                    concat_crescente_filtrado = pd.concat([concat_crescente_filtrado, df_crescente_filtrado], ignore_index = True)
                    lista_crescente_filtrado.append(df_crescente_filtrado)
                    concat_crescente_absoluto = pd.concat([concat_crescente_absoluto, df_crescente_absoluto], ignore_index = True)
                    lista_crescente_absoluto.append(df_crescente_absoluto)
                    lista_cabecalho_crescente.append(cabecalho_crescente)
                    lista_km_crescente.append(km_crescente)
                    lista_estacoes_crescente.append(estacoes_crescente)
                    lista_secoes_km.append(secoes_km)
                # Exporta inventário no modelo pavesys e dersp
                if pavesys or pavesys_faixa:
                    AnexoAPavesys(path, modelo_path_anexoA, pista.split(".")[0] + '.xlsx', lista_cabecalho_crescente[0], concat_crescente_absoluto)
                if dersp:
                    AnexoADERSP(path, modelo_path_anexoA_SP, pista.split(".")[0] + '.xlsx', lista_cabecalho_crescente[0], concat_crescente_absoluto)
            
                # Exporta fixa no modelo pavesys e dersp
                Valores_Ficha = {}
                contador_ficha = 0
                for adc in range(len(lista_secoes_km)):
                    for secao in range(len(lista_secoes_km[adc]) - 1):
                        df = lista_crescente_filtrado[adc][(lista_crescente_filtrado[adc]['Início'] >= lista_secoes_km[adc][secao]) & (lista_crescente_filtrado[adc]['Início'] < lista_secoes_km[adc][secao + 1])]
                        num_estacoes = df['Início'].count()
                        df_abs = lista_crescente_absoluto[adc][(lista_crescente_absoluto[adc]['Início'] >= lista_secoes_km[adc][secao]) & (lista_crescente_absoluto[adc]['Início'] < lista_secoes_km[adc][secao + 1])]
                        Valores_Ficha[contador_ficha] = FichaIGG(df, df_abs, num_estacoes, [lista_secoes_km[adc][secao], lista_secoes_km[adc][secao + 1]])
                        contador_ficha = contador_ficha + 1
                if pavesys or pavesys_faixa:
                    AnexoBPavesys(path, modelo_path_anexoB, pista.split(".")[0] + '.xlsx', Valores_Ficha, contador_ficha, cabecalho_crescente)
                if dersp:
                    AnexoBDERSP(path, modelo_path_anexoB_SP, pista.split(".")[0] + '.xlsx', Valores_Ficha, contador_ficha, cabecalho_crescente, modelo_path_unifilar, concat_crescente_absoluto)
                if antt:
                    AnexoANTT(path, modelo_path_simples, pista.split(".")[0] + '.xlsx', cabecalho_crescente, concat_crescente_absoluto, Valores_Ficha, contador_ficha)
                    
                # Armazenando valores para P15
                if artesp:
                    df_p15 = pd.concat([df_p15, concat_crescente_absoluto], ignore_index = True)
                    df_p15.to_excel(os.path.join(path, "copia de seguranca - P15.xlsx"))
                    GeraP21(path, pista.split(".")[0], Valores_Ficha, contador_ficha, True)

    # Exportando planilha P15
    if artesp:
        df_p15.sort_values(by = ["Rodovia", "Sentido", "km referencia", "Nome ramo", "Faixa"], ascending = [True, True, True, True, True], inplace = True, ignore_index = True)
        GeraP15(path, modelo_path_p15, 'P15.xlsx', df_p15)
        os.remove(os.path.join(path, "copia de seguranca - P15.xlsx"))

    print("Processamento de IGG finalizado!")



#print('Encontrando diretório dos arquivo: ')
#path = r'C:\Users\Pavesys - MAQ70\Desktop\Nova pasta'

# Modelos desejados de exportação
#antt = True
#artesp = True
#dersp = True
#pavesys = True
#pavesys_faixa = True
#segmento_homogeneo = 1

#IGG(path, antt, artesp, dersp, pavesys, pavesys_faixa, segmento_homogeneo)



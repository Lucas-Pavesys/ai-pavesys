# ------ IMPORTAÇÕES ------
import os
import pandas as pd
import numpy as np
import openpyxl as op
import openpyxl.styles as ops
from math import trunc
import ler_arquivo_V01 as leitor


# ------ FUNÇÕES ------
def FiltroDeIndice(indice):
    ini_fim = []
    if len(indice) != 0:
        ini_fim = [indice[0]]
        for valor in range(len(indice) - 1):
            if indice[valor + 1 ] - indice[valor] != 1:
                ini_fim.append(indice[valor])
                ini_fim.append(indice[valor + 1])
        ini_fim.append(indice[-1])
    else:
        pass
    return ini_fim


def FormataLinha(patologia, indice_defeitos_filtrado, lista_kmini, lista_kmfim, lista_latitude, lista_longitude, dict_descricao, posicao, largura_defeito):
    lista_final = []
    i = 0
    if patologia == 'TI':
        largura_defeito = 0.15
    while i < len(indice_defeitos_filtrado) - 1:
        indice_ini = indice_defeitos_filtrado[i]
        indice_fim = indice_defeitos_filtrado[i + 1]
        tuple_temp = (lista_kmini[indice_ini], 
                    lista_kmfim[indice_fim], 
                    lista_latitude[indice_ini], 
                    lista_longitude[indice_ini], 
                    lista_latitude[indice_fim], 
                    lista_longitude[indice_fim], 
                    dict_descricao[patologia][0], 
                    dict_descricao[patologia][1], 
                    dict_descricao[patologia][2], 
                    round(abs(lista_kmini[indice_ini] - lista_kmfim[indice_fim]) * 1000, 3), 
                    posicao, 
                    round(abs(lista_kmini[indice_ini] - lista_kmfim[indice_fim]) * 1000 * largura_defeito, 3))
        lista_final.append(tuple_temp)
        i = i + 2
    return lista_final


def ConcatenaDf(coluna, *args): # RECEBE APENAS DATAFRAME
    # Função auxiliar da: PadraoLVC
    colunas = pd.concat(args, axis=1)
    matriz_colunas = np.array(colunas)
    lista_concatenada = []
    for a in range(len(matriz_colunas)):
        if 'x' in list(matriz_colunas[a]):
            lista_concatenada.append('x')
        else:
            lista_concatenada.append(None)

    df_concatenado = pd.DataFrame(lista_concatenada)
    df_concatenado.columns = coluna
    return df_concatenado


def ExportaExcel(df, rodovia, faixa, pista, inicio, fim, largura, data, path, nome):
    # Criando arquivo excel para exportar
    wb = op.Workbook()
    sheet = wb.active
    sheet.title = '_'

    # Criando cabeçalho
    sheet['A1'] = 'LISTA DE DEFEITOS DA SUPERFÍCIE DO PAVIMENTO (LVD-Vídeo)'
    sheet['A3'] = 'Rodovia:'
    sheet['B3'] = rodovia
    sheet['A4'] = 'Faixa:'
    sheet['B4'] = faixa
    sheet['D3'] = 'Pista:'
    sheet['E3'] = pista
    sheet['D4'] = 'Trecho:'
    sheet['E4'] = "km "+str(f'{inicio:.3f}').replace(".","+")+" ao km "+str(f'{fim:.3f}').replace(".","+")
    sheet['H3'] = 'Largura Faixa (m):'
    sheet['I3'] = largura
    sheet['H4'] = 'Data:'
    sheet['I4'] = data

    sheet['A6'] = 'Localização (km)'
    sheet.merge_cells('A6:B6')
    sheet['C6'] = 'Georref (Início)'
    sheet.merge_cells('C6:D6')
    sheet['E6'] = 'Georref (Final)'
    sheet.merge_cells('E6:F6')
    sheet['G6'] = 'Sigla'
    sheet.merge_cells('G6:G7')
    sheet['H6'] = 'Descrição'
    sheet.merge_cells('H6:H7')
    sheet['I6'] = 'Severidade'
    sheet.merge_cells('I6:I7')
    sheet['J6'] = 'Comprimento (m)'
    sheet.merge_cells('J6:J7')
    sheet['K6'] = 'Localização'
    sheet.merge_cells('K6:K7')
    sheet['L6'] = 'Área (m²)'
    sheet.merge_cells('L6:L7')
    sheet['A7'] = 'Inicial'
    sheet['B7'] = 'Final'
    sheet['C7'] = 'Latitude'
    sheet['D7'] = 'Lontitude'
    sheet['E7'] = 'Latitude'
    sheet['F7'] = 'Lontitude'

    # Formatando título principal
    sheet.merge_cells('A1:L1')
    titulo = ops.Font(name = 'Calibri', size = 12, bold = True)
    centro = ops.Alignment(horizontal = "center", vertical = "center")
    sheet['A1'].font = titulo
    sheet['A1'].alignment = centro

    # Formatando restante do cabeçalho
    subtitulo = ops.Font(name = 'Calibri', size = 11, bold = True)
    esquerda = ops.Alignment(horizontal = "left", vertical = "center")
    lista = ['A3', 'A4', 'D3', 'D4', 'H3', 'H4']
    for cel in lista:
        sheet[cel].font = subtitulo
        sheet[cel].alignment = esquerda
    texto = ops.Font(name = 'Calibri', size = 11)
    lista = ['B3', 'B4', 'E3', 'E4', 'I3', 'I4']
    for cel in lista:
        sheet[cel].font = texto
        sheet[cel].alignment = esquerda
    sheet['I4'].number_format = "mmm-yy"

    # Formatando cabeçaho da tabela
    cabecalho = ops.Font(name = 'Calibri', size = 10, bold = True)
    preenchimento = ops.PatternFill(patternType='solid', fgColor='E0E0E0')
    lista = ['A6', 'C6', 'E6', 'G6', 'H6', 'I6', 'J6', 'K6', 'L6', 'A7', 'B7', 'C7', 'D7', 'E7', 'F7']
    for cel in lista:
        sheet[cel].font = cabecalho
        sheet[cel].alignment = centro
        sheet[cel].fill = preenchimento

    # Formatando largura de colunas
    sheet.column_dimensions['A'].width = 9
    sheet.column_dimensions['B'].width = 9
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 4
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 10
    sheet.column_dimensions['J'].width = 4
    sheet.column_dimensions['K'].width = 10
    sheet.column_dimensions['L'].width = 10

    # Organizando dados do dataframe
    corpo = ops.Font(name = 'Calibri', size = 10)
    for linha in range(len(df)):
        for coluna in range(len(df.columns)):
            sheet.cell(row = linha + 8, column = coluna + 1).value = df.iloc[linha, coluna]
            sheet.cell(row = linha + 8, column = coluna + 1).font = corpo
            sheet.cell(row = linha + 8, column = coluna + 1).alignment = centro

    # Salvando arquivo excel
    nomeLVD = os.path.join(path, nome)
    wb.save(nomeLVD)


# Gera o df_patologias
def OrganizaPatologias(file_loc):
    df = pd.read_excel(file_loc, header=7, usecols="A:BC")
    for name in list(df.columns.values.tolist()):
        if "TLC.FC-23.F" in name:
            teclas_var = True
            break
        else:
            teclas_var = False
    if teclas_var == True:
        df = pd.read_excel(file_loc, header=7)
        df_ini_fim = pd.concat([df['Início'], df['Fim']], axis=1)
        df_ini_fim.columns = ['Início', 'Fim']
        df_fi_be = ConcatenaDf(['FI.FC-1.BE'], df['Fi.FC-1.BE'], df['J1.FC-1.BE'])
        df_fi_atre = ConcatenaDf(['FI.FC-1.ATRE'], df['Fi.FC-1.ATRE'], df['J1.FC-1.ATRE'])
        df_fi_f = ConcatenaDf(['FI.FC-1.F'], df['Fi.FC-1.F'], df['J1.FC-1.F'])
        df_fi_atrd = ConcatenaDf(['FI.FC-1.ATRD'], df['Fi.FC-1.ATRD'], df['J1.FC-1.ATRD'])
        df_fi_bd = ConcatenaDf(['FI.FC-1.BD'], df['Fi.FC-1.BD'], df['J1.FC-1.BD'])
        df_j_je = pd.concat([df['J.FC-2.BE'], df['J.FC-2.ATRE'], df['J.FC-2.F'], df['J.FC-2.ATRD'], df['J.FC-2.BD'], df['JE.FC-3.BE'], df['JE.FC-3.ATRE'], df['JE.FC-3.F'], df['JE.FC-3.ATRD'], df['JE.FC-3.BD']], axis=1)
        df_ti_be = ConcatenaDf(['TI.FC-23.BE'], df['TTC.FC-23.BE'], df['TTL.FC-23.BE'], df['TLC.FC-23.BE'], df['TLL.FC-23.BE'])
        df_ti_atre = ConcatenaDf(['TI.FC-23.ATRE'], df['TTC.FC-23.ATRE'], df['TTL.FC-23.ATRE'], df['TLC.FC-23.ATRE'], df['TLL.FC-23.ATRE'])
        df_ti_f = ConcatenaDf(['TI.FC-23.F'], df['TTC.FC-23.F'], df['TTL.FC-23.F'], df['TLC.FC-23.F'], df['TLL.FC-23.F'])
        df_ti_atrd = ConcatenaDf(['TI.FC-23.ATRD'], df['TTC.FC-23.ATRD'], df['TTL.FC-23.ATRD'], df['TLC.FC-23.ATRD'], df['TLL.FC-23.ATRD'])
        df_ti_bd = ConcatenaDf(['TI.FC-23.BD'], df['TTC.FC-23.BD'], df['TTL.FC-23.BD'], df['TLC.FC-23.BD'], df['TLL.FC-23.BD'])
        df_afo_be = ConcatenaDf(['Afund/Ond.BE'], df['ALP-23.BE'], df['ALC-23.BE'], df['ATP-23.BE'], df['ATC-23.BE'], df['OND.BE'])
        df_afo_atre = ConcatenaDf(['Afund/Ond.ATRE'], df['ALP-23.ATRE'], df['ALC-23.ATRE'], df['ATP-23.ATRE'], df['ATC-23.ATRE'], df['OND.ATRE'])
        df_afo_f = ConcatenaDf(['Afund/Ond.F'], df['ALP-23.F'], df['ALC-23.F'], df['ATP-23.F'], df['ATC-23.F'], df['OND.F'])
        df_afo_atrd = ConcatenaDf(['Afund/Ond.ATRD'], df['ALP-23.ATRD'], df['ALC-23.ATRD'], df['ATP-23.ATRD'], df['ATC-23.ATRD'], df['OND.ATRD'])
        df_afo_bd = ConcatenaDf(['Afund/Ond.BD'], df['ALP-23.BD'], df['ALC-23.BD'], df['ATP-23.BD'], df['ATC-23.BD'], df['OND.BD'])
        df_p = pd.concat([df['Panela.BE.A'], df['Panela.BE.M'], df['Panela.BE.B'], df['Panela.ATRE.A'], df['Panela.ATRE.M'], df['Panela.ATRBE.B'], df['Panela.F.A'], 
                          df['Panela.F.M'], df['Panela.F.B'], df['Panela.ATRD.A'], df['Panela.ATRD.M'], df['Panela.ATRD.B'], df['Panela.BD.A'], df['Panela.BD.M'], df['Panela.BD.B']], axis=1)
        df_p.columns = ['Panela.BE.A', 'Panela.BE.M', 'Panela.BE.B', 'Panela.ATRE.A', 'Panela.ATRE.M', 'Panela.ATRE.B', 'Panela.F.A', 'Panela.F.M', 'Panela.F.B', 'Panela.ATRD.A', 
                        'Panela.ATRD.M', 'Panela.ATRD.B', 'Panela.BD.A', 'Panela.BD.M', 'Panela.BD.B']
        df_ex = pd.concat([df['Exsudação.BE'], df['Exsudação.ATRE'], df['Exsudação.F'], df['Exsudação.ATRD'], df['Exsudação.BD']], axis=1)
        df_r = pd.concat([df['Remendo.BE'], df['Remendo.ATRE'], df['Remendo.F'], df['Remendo.ATRD'], df['Remendo.BD']], axis=1)
        df_coordenadas = pd.concat([df['Observação'], df['Latitude'], df['Longitude']], axis=1)
        df_coordenadas.columns = ['Observação', 'Latitude', 'Longitude']
        df_patologias = pd.concat([df_ini_fim, df_fi_be, df_fi_atre, df_fi_f, df_fi_atrd, df_fi_bd, df_j_je, df_ti_be, df_ti_atre, df_ti_f, 
                                   df_ti_atrd, df_ti_bd, df_r, df_p, df_ex, df_afo_be, df_afo_atre, df_afo_f, df_afo_atrd, df_afo_bd, df_coordenadas], axis=1)
        return df_patologias
    
    else:
        return df


# Gera o df organizado para exportar no excel (df_final_ordenado)
def GeraDF(df):
    # Listas de km e coordenadas
    lista_kmini = df['Início'].tolist()
    lista_kmfim = df['Fim'].tolist()
    lista_latitude = df['Latitude'].tolist()
    lista_longitude = df['Longitude'].tolist()

    # Df separado com patologias
    df_patologias = df.drop(['index', 'Início', 'Fim', 'Observação', 'Latitude', 'Longitude'], axis=1)

    # Lista de defeitos
    cabecalho_defeitos = df_patologias.columns.values.tolist()

    dict_descricao = {"FI": ("Fi", "Fissuras", "FC-1"),
                    "J": ("J", "Trincas Interligadas", "FC-2"),
                    "JE": ("JE", "Tricas Interligadas", "FC-3"),
                    "TI": ("TI", "Trincas Isoladas", "FC-2/3"),
                    "Panela": ("P", "Panela", "-"),
                    "Exsudação": ("EX", "Exsudação", "-"),
                    "Remendo": ("R", "Remendos", "-"),
                    "Afund/Ond": ("Afund/Ond", "Afundamento ou Ondulação", "-")}

    dict_posicao = {"BE": [0.4, "Bordo Esq."], "ATRE": [0.8, "ATR Esq."], "F": [3.6, "Faixa"], "ATRD": [0.8, "ATR Dir."], "BD": [0.4, "Bordo Dir."]}

    # Identificando crescente e decrescente
    if lista_kmini[0] - lista_kmfim[-1] < 0:
        crescente = True
    else:
        crescente = False

    # Pegando índice das patologias
    indice_defeitos_filtrado = []
    for defeito in cabecalho_defeitos:
        lista_defeitos = df_patologias[defeito].tolist()
        indice_defeitos = df_patologias.index[df_patologias[defeito] == "x"].tolist()
        indice_defeitos_filtrado.append(FiltroDeIndice(indice_defeitos))
    
    # Calulando áreas e comprimentos - Configurando linhas
    lista_LVD = []
    patologias_keys = []

    for j in range(len(cabecalho_defeitos)):
        patologias_keys.append(cabecalho_defeitos[j].split(".")[0].split("-")[0])

    for j in range(len(patologias_keys)):
        if cabecalho_defeitos[j].split(".")[-1] in dict_posicao.keys():
            lista_LVD = lista_LVD + FormataLinha(patologias_keys[j], indice_defeitos_filtrado[j], lista_kmini, lista_kmfim, lista_latitude, lista_longitude, dict_descricao, 
                                                dict_posicao[cabecalho_defeitos[j].split(".")[-1]][1], dict_posicao[cabecalho_defeitos[j].split(".")[-1]][0])
        elif cabecalho_defeitos[j].split(".")[1] in dict_posicao.keys():
            lista_LVD = lista_LVD + FormataLinha(patologias_keys[j], indice_defeitos_filtrado[j], lista_kmini, lista_kmfim, lista_latitude, lista_longitude, dict_descricao, 
                                                dict_posicao[cabecalho_defeitos[j].split(".")[1]][1], dict_posicao[cabecalho_defeitos[j].split(".")[1]][0])

    # Organizando as linhas 
    index = ['Inicial', 'Final', 'Latitude ini', 'Longitude ini', 'Latitude fim', 'Longitude fim', 'Sigla', 'Descrição', 'Severidade', 'Comprimento (m)', 'Localização', 'Área (m²)']
    df_final = pd.DataFrame(lista_LVD, columns=index)

    if crescente is True:
        df_final_ordenado = df_final.sort_values('Inicial')
    else:
        df_final_ordenado = df_final.sort_values('Inicial', ascending=False)

    return df_final_ordenado, lista_kmini, lista_kmfim


def ExportaLVDPavesys(file_loc, planilha, lista_kmini, lista_kmfim, df_final_ordenado, path):
    # Exportando excel
    dadosgerais = pd.read_excel(file_loc, usecols = "B", nrows = 5)
    rodovia = list(dadosgerais.columns)[0]
    faixa = dadosgerais.iloc[1, 0]
    pista = dadosgerais.iloc[0, 0]
    inicio = float(dadosgerais.iloc[2, 0])
    fim = float(dadosgerais.iloc[3, 0])
    largura = 3.6
    data = dadosgerais.iloc[4, 0]

    nome = 'LVD-Pavesys_' + planilha.split(".")[0] + '.xlsx'
    ExportaExcel(df_final_ordenado, rodovia, faixa, pista, inicio, fim, largura, data, path, nome)
    print("... pronto para uso.\n")


# Função para chamar no botão do software
def BotaoLVDPavesys(path):
    # Organizando arquivos para processamento
    """
    dc_pista_simples, dc_pista_dupla_cresc, dc_pista_dupla_decresc, dc_pista_adc_cres, dc_pista_adc_decres, dc_pista_ramo, atr_pista_simples, atr_pista_dupla_cresc, atr_pista_dupla_decresc, atr_pista_adc_cres, atr_pista_adc_decres, atr_pista_ramo = leitor.dicionario_arquivos(path)
    dict_trechos = [dc_pista_simples, dc_pista_dupla_cresc, dc_pista_dupla_decresc, dc_pista_adc_cres, dc_pista_adc_decres, dc_pista_ramo]
    """
    trechos_LVD = []
    for filename in os.listdir(path): # filtrando arquivos excel
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if not filename.split(".")[0].endswith("_ATR"):
                trechos_LVD.append(filename)

    for planilha in trechos_LVD:
        # Arquivo a ser processado
        print("Iniciando o processamento de " + planilha)
        file_loc = os.path.join(path, planilha)
        df_patologias = OrganizaPatologias(file_loc)

        secoes = []
        for valor in df_patologias['Início']:
            if valor*1000 % 1000 == 0:
                secoes.append(valor)
        if len(secoes) == 0:
            secoes.append(df_patologias['Início'].tolist()[0])
        if secoes[-1] == df_patologias['Início'].tolist()[-1]:
            pass
        else:
            secoes.append(df_patologias['Início'].tolist()[-1])
        if secoes[0] == df_patologias['Início'].tolist()[0]:
            pass
        else:
            secoes.insert(0, df_patologias['Início'].tolist()[0])
                
        if secoes[0] - secoes[1] < 0:
            crescente = True
        else:
            crescente = False

        # Exportando planilha avaliada a cada 20m
        LVC20m(df_patologias, path, planilha)

        df_final = pd.DataFrame()
        lista_kmini = []
        lista_kmfim = []
        secao=0
        if crescente:
            while secao < len(secoes)-1:
                df = df_patologias[(df_patologias['Início'] >= secoes[secao]) & (df_patologias['Início'] < secoes[secao + 1])]
                df_final_ordenado, kmini, kmfim = GeraDF(df.reset_index())
                df_final = pd.concat([df_final, df_final_ordenado])
                lista_kmini = lista_kmini + kmini
                lista_kmfim = lista_kmfim + kmfim
                secao = secao + 1
        else:
            while secao < len(secoes)-1:
                df = df_patologias[(df_patologias['Início'] <= secoes[secao]) & (df_patologias['Início'] > secoes[secao + 1])]
                df_final_ordenado, kmini, kmfim = GeraDF(df.reset_index())
                df_final = pd.concat([df_final, df_final_ordenado])
                lista_kmini = lista_kmini + kmini
                lista_kmfim = lista_kmfim + kmfim
                secao = secao + 1
        ExportaLVDPavesys(file_loc, planilha, lista_kmini, lista_kmfim, df_final, path)
    
    print("LVD Pavesys finalizado!")


# Função para código do IGG no relatório versão artesp | Espaçamento em km
def AreaTrincada(file_loc, secoes, ini, fim):
    # Arquivo a ser processado
    df_patologias = OrganizaPatologias(file_loc)
    
    # Seções para calculo de área de defeitos
    if secoes[0] != ini:
        secoes.insert(0, ini)
    if secoes[-1] != fim:
        secoes.append(fim)
    
    if ini - fim < 0:
        crescente = True
    else:
        crescente = False
    df_patologias.drop(['FI.FC-1.BE', 'FI.FC-1.ATRE', 'FI.FC-1.F', 'FI.FC-1.ATRD', 'FI.FC-1.BD', 'TI.FC-23.BE', 'TI.FC-23.ATRE', 'TI.FC-23.F', 'TI.FC-23.ATRD', 'TI.FC-23.BD', 'Remendo.BE', 
                         'Remendo.ATRE', 'Remendo.F', 'Remendo.ATRD', 'Remendo.BD', 'Exsudação.BE', 'Exsudação.ATRE', 'Exsudação.F', 'Exsudação.ATRD', 'Exsudação.BD'], axis='columns', inplace=True)
    
    df_final = pd.DataFrame()
    secao=0
    if crescente:
        df_patologias = df_patologias[(df_patologias['Início'] >= ini) & (df_patologias['Início'] < fim)]
        while secao < len(secoes)-1:
            df = df_patologias[(df_patologias['Início'] >= secoes[secao]) & (df_patologias['Início'] < secoes[secao + 1])]
            df_final_ordenado, kmini, kmfim = GeraDF(df.reset_index())
            df_final = pd.concat([df_final, df_final_ordenado])
            secao = secao + 1
    else:
        df_patologias = df_patologias[(df_patologias['Início'] <= ini) & (df_patologias['Início'] > fim)]
        while secao < len(secoes)-1:
            df = df_patologias[(df_patologias['Início'] <= secoes[secao]) & (df_patologias['Início'] > secoes[secao + 1])]
            df_final_ordenado, kmini, kmfim = GeraDF(df.reset_index())
            df_final = pd.concat([df_final, df_final_ordenado])
            secao = secao + 1

    return df_final


def SeparaDF (df):
    # Criar uma lista de dataframes separados
    dfs_separados = []
    df_atual = pd.DataFrame()

    # Iterar sobre as linhas do dataframe original
    for index, row in df.iterrows():
        # Verificar se a linha está em branco
        if row.isnull().all():
            # Adicionar o dataframe atual à lista de dataframes separados
            if not df_atual.empty:
                dfs_separados.append(df_atual)
                df_atual = pd.DataFrame()
        else:
            # Adicionar a linha atual ao dataframe atual
            df_atual = pd.concat([df_atual, df.loc[[index]]])

    # Adicionar o último dataframe atual à lista de dataframes separados
    if not df_atual.empty:
        dfs_separados.append(df_atual)
    return dfs_separados


# Função para concatenar a planilha 1m em 20m
def LVC20m(df, path, planilha):
    lista_dfs = SeparaDF(df)

    df_concatenado = pd.DataFrame()
    for df in lista_dfs:
        secoes = []
        km = df['Início'].tolist()
        crescente = True if km[0] < km[-1] else False

        # Separando seções para concatenar
        for valor in km:
            if round(valor*1000, 3) % 20 == 0:
                secoes.append(valor)
        if len(secoes) == 0:
            secoes.append(km[0])
        if secoes[0] != km[0]:
            secoes.insert(0, km[0])
        if secoes[-1] != df['Fim'].tolist()[-1]:
            secoes.append(df['Fim'].tolist()[-1])
        
        for trecho in range(len(secoes)-1):
            if crescente:
                df_temp = df[(df['Início'] >= secoes[trecho]) & (df['Início'] < secoes[trecho + 1])]
            else:
                df_temp = df[(df['Início'] <= secoes[trecho]) & (df['Início'] > secoes[trecho + 1])]

            for coluna in df_temp:
                if 'x' in df_temp[coluna].tolist():
                    df_temp[coluna] = 'x'
            
            df_temp = df_temp[(df_temp['Início']*1000 == secoes[trecho]*1000)]
            df_temp['Início'] = secoes[trecho]
            df_temp['Fim'] = secoes[trecho + 1]

            df_concatenado = pd.concat([df_concatenado, df_temp])

    df_concatenado.to_excel(os.path.join(path, "LVC 20m_" + planilha.split(".")[0] + ".xlsx"), index=False)


# -------------------- CÓDIGO --------------------

# print('Encontrando diretório dos arquivo: ')
# path = os.getcwd()
# path = r'C:\Users\Pavesys - MAQ70\OneDrive - Pavesys Engenharia\Contratos Equipe Gabriel\_CLN (2021.20)\SAUIPE - CLN\SAUIPE-CRES-1-0-1'

# BotaoLVDPavesys(path)


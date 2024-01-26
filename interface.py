import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import os
import openpyxl
from pre_processamento import pre_processo as pp
from IGG_V03 import IGG
from LVD_Pavesys_V03 import BotaoLVDPavesys
import pandas as pd
from numpy import nan
from lvc_dnit import lvc_dnit as ld
import shutil
from datetime import date

for file in os.listdir(os.getcwd()):
    if file.__contains__("LVC 1m_MODELO.xlsx"):
        modelo_path = os.path.join(os.getcwd(), file)
    if file.__contains__("MODELO-LVC DNIT.xlsx"):
        lvcdnit_path = os.path.join(os.getcwd(), file)
    if file.endswith(".tcl"):
        theme_path = os.path.join(os.getcwd(), file)
    if file.endswith(".ico"):
        icon_path = os.path.join(os.getcwd(), file)
    if file.__contains__("ATR"):
        atr_modelo = os.path.join(os.getcwd(), file)

class Fctn_Processamento:
    def exit_application():
        # Sai da aplicação
        if messagebox.askyesno("Sair", "Tem certeza que gostaria de sair?"):
            root.destroy()

    def select_directory():
        # Função que seleciona o diretório e insere na list_box e Directory Entry
        directory = filedialog.askdirectory()
        if directory and os.path.isdir(directory):
            files = os.listdir(directory)

            directory_entry.delete(0, tk.END)
            directory_entry.insert(0, directory)
            files_listbox.delete(*files_listbox.get_children())

            directory_entry_lvc.delete(0, tk.END)
            directory_entry_lvc.insert(0, directory)
            files_listbox_lvc.delete(0, tk.END)

            directory_entry_igg.delete(0, tk.END)
            directory_entry_igg.insert(0, directory)
            files_listbox_igg.delete(0, tk.END)

            directory_entry_lvd.delete(0, tk.END)
            directory_entry_lvd.insert(0, directory)
            files_listbox_lvd.delete(0, tk.END)

            directory_entry_lvdd.delete(0, tk.END)
            directory_entry_lvdd.insert(0, directory)
            files_listbox_lvdd.delete(0, tk.END)
            for file in files:
                if (
                    not file.startswith("PP")
                    and file.endswith(".xlsx")
                    or file.endswith(".xls")
                ):
                    files_listbox.insert("", "end", text=file)

                if  file.startswith("PP"):

                    files_listbox_lvc.insert(tk.END, file)
                    files_listbox_igg.insert(tk.END, file)
                    files_listbox_lvd.insert(tk.END, file)
                    files_listbox_lvdd.insert(tk.END, file)

    def on_enter_pressed(event):
        page2.focus()  # Shift focus to another widget

    def refresh_listboxes(directory):
        files = os.listdir(directory)

        files_listbox.delete(*files_listbox.get_children())
        files_listbox_lvc.delete(0, tk.END)
        files_listbox_igg.delete(0, tk.END)
        files_listbox_lvd.delete(0, tk.END)

        for file in files:
            if not file.startswith("PP") and (
                file.endswith(".xlsx") or file.endswith(".xls")
            ):
                files_listbox.insert("", "end", text=file)
            if (
                file.endswith(".xlsx")
                or file.endswith(".xls")
                and file.startswith("PP")
            ):
                files_listbox_lvc.insert(tk.END, file)
                files_listbox_igg.insert(tk.END, file)
                files_listbox_lvd.insert(tk.END, file)

    def pp_save(modelo_path, file_path, df, sheet_name, start_col, start_row):
        book = openpyxl.load_workbook(modelo_path)
        sheet = book[sheet_name]
        current_row = start_row

        info_cabeçalho = str(name_spinbox.get()).split("_")

        try:

            cabeçalho_lista = [info_cabeçalho[0],   # Rodovia
                                info_cabeçalho[1],  # Pista
                                info_cabeçalho[-1], # Faixa
                                ini_entry.get(),    # Início
                                fim_entry.get(),    # Fim
                                date.today()]       # Data
        except: 

            cabeçalho_lista = [info_cabeçalho[0],   # Rodovia
                                info_cabeçalho[1],  # Pista
                                info_cabeçalho[-1], # Faixa
                                ini_entry.get(),    # Início
                                fim_entry.get(),    # Fim
                                date.today()]       # Data
        
        for i, value in enumerate(cabeçalho_lista):
            sheet.cell(row= 1 + i, column=2, value=value)
            
        for col, header in enumerate(df.columns):
            sheet.cell(row=current_row, column=start_col + col).value = header
        current_row += 1

        x = 1

        for _, row in df.iterrows():
            pp_progress["value"] = x
            percent = round((x / len(df)) * 100, 2)
            pp_var.set(f"{percent} %")

            root.update()
            page1.update()
            page1.update_idletasks()

            for col, value in enumerate(row, start=start_col):
                sheet.cell(row=current_row, column=col, value=value)
            current_row += 1
            x += 1

        save_var.set("Salvando o Arquivo...")
        root.update()
        page1.update()
        book.save(file_path)

    def lvc_save(modelo_path, file_path, df, sheet_name, start_col, start_row):
        book = openpyxl.load_workbook(modelo_path)
        sheet = book[sheet_name]
        current_row = start_row
        
        x = 1

        for _, row in df.iterrows():
            lvc_progress["value"] = x
            percent = round((x / len(df)) * 100, 2)
            lvc_var.set(f"{percent} %")

            root.update()
            page2.update()
            page2.update_idletasks()

            for col, value in enumerate(row, start=start_col):
                sheet.cell(row=current_row, column=col, value=value)
            current_row += 1
            x += 1

        root.update()
        page2.update()

        book.save(file_path)

    def run_pre_processamento():
        src_path = directory_entry.get()
        xl_name = str("PP_" + name_spinbox.get() + ".xlsx")

        dst_path = os.path.join(src_path, xl_name)

        file_list = process_listbox.get(0, tk.END)
        ini_var = float(ini_entry.get())
        fim_var = float(fim_entry.get())
        tipo = tipo_var.get()

        pp_parameters = pp(src_path, file_list, ini_var, fim_var, tipo)
        pp_df = pp_parameters.run_pp()

        pp_lenght = len(pp_df)
        pp_progress.configure(maximum=pp_lenght)

        Fctn_Processamento.pp_save(modelo_path, dst_path, pp_df, "_", 1, 8)
        pp_progress["value"] = 0
        pp_var.set("")
        save_var.set("")

        page1.update_idletasks()
        messagebox.showinfo("Aviso", "Planilha Pre-Processada exportada!")

    def run_pp_list():
        src_path = directory_entry.get()
        xl_name = str("PP_" + list_label.get() + ".xlsx")
        dst_path = os.path.join(src_path, xl_name)

        file_list = process_listbox.get(0, tk.END)

        main_path = os.getcwd()
        list_path = os.path.join(main_path, "cut_list.txt")
        df = pd.read_table(list_path, header=None, decimal=",")
        df.columns = ["Início", "Fim"]
        tipo = tipo_var.get()

        result_dfs = []

        for i in range(len(df)):
            ini_var = float((df["Início"].iloc[i]))
            fim_var = float(df["Fim"].iloc[i])
            pp_parameters = pp(src_path, file_list, ini_var, fim_var, tipo)
            pp_df = pp_parameters.run_pp()

            result_dfs.append(pp_df)
            empty_row = pd.DataFrame(nan, index=[0], columns=pp_df.columns)
            result_dfs.append(empty_row)

        final_result = pd.concat(result_dfs, ignore_index=True)

        pp_lenght = len(final_result)
        pp_progress.configure(maximum=pp_lenght)

        Fctn_Processamento.pp_save(modelo_path, dst_path, final_result, "_", 1, 8)
        pp_progress["value"] = 0
        pp_var.set("")
        save_var.set("")

        page1.update_idletasks()
        new_window.end()
        messagebox.showinfo("Aviso", "Planilha Pre-Processada exportada!")

    def run_lvc_dnit():

        src_path = directory_entry.get()
        intervalo = int(sep_lvcdnit.get())
        
        tipo = tipo_var.get()

        file_list = files_listbox_lvc.get(0, tk.END)
        file_name = (file_list[0].split("_"))[1]
        if len(file_list) == 0:
            messagebox.showinfo("Aviso", "Pasta Vazia!")
        else:
            ttk_labels = []
            txt_heigth = 40
            ini_do_processo = ttk.Label(page2, text=" ======================== INÍCIO =======================", background="gray28")
            ini_do_processo.place(x=410, y=20)
            ttk_labels.append(ini_do_processo)
            txt_trecho = ttk.Label(page2, text=("> " + file_name), background="gray28")
            txt_trecho.place(x=420, y=txt_heigth)
            ttk_labels.append(txt_trecho)
            txt_heigth += 20
            root.update_idletasks()
            page2.update_idletasks()

            dst_path = os.path.join(src_path, str("LVC_" + file_name + ".xlsx"))
            shutil.copy(lvcdnit_path, dst_path)
            for file in file_list:
                if file.startswith("PP") and "ATR" not in file:
                    lvc_path = os.path.join(src_path, file)
                    file = file.replace(".xlsx", "").replace(
                        ".xls", ""
                    )  # Fix: you need to assign the result to 'file'
                    file = file.split("_")
                    if file[-1] == "1":
                        # FX1
                        if file[3] == "C":
                            sheet_name = "Faixa 1 Crescente"
                            fx1_c = (f"  > Processando: {sheet_name}...")
                            lfx1_c = ttk.Label(page2, text=fx1_c, background="gray28")
                            lfx1_c.place(x=430, y=txt_heigth)
                            lvc_var.set("")
                            ttk_labels.append(lfx1_c)
                        else:
                            sheet_name = "Faixa 1 Decrescente"
                            fx1_d = (f"  > Processando: {sheet_name}...")
                            lfx1_d = ttk.Label(page2, text=fx1_d, background="gray28")
                            lfx1_d.place(x=430, y=txt_heigth)
                            lvc_var.set("")         
                            ttk_labels.append(lfx1_d)                  
                    elif file[-1] == "2":
                        # FX2
                        if file[3] == "C":
                            sheet_name = "Faixa 2 Crescente"
                            fx2_c = (f"  > Processando: {sheet_name}...")
                            lfx2_c = ttk.Label(page2, text=fx2_c, background="gray28")
                            lfx2_c.place(x=430, y=txt_heigth)
                            lvc_var.set("")
                            ttk_labels.append(lfx2_c)
                        else:
                            sheet_name = "Faixa 2 Decrescente"
                            fx2_d = (f"  > Processando: {sheet_name}...")
                            lfx2_d = ttk.Label(page2, text=fx2_d, background="gray28")
                            lfx2_d.place(x=430, y=txt_heigth)
                            lvc_var.set("")
                            ttk_labels.append(lfx2_d)
                    elif file[-1] == "3" in file[-1]:
                        # FX3 ou ADC
                        if file[3] == "C":
                            sheet_name = "Faixa 3 Crescente"
                            fx3_c = (f"  > Processando: {sheet_name}...")
                            lfx3_c = ttk.Label(page2, text=fx3_c, background="gray28")
                            lfx3_c.place(x=430, y=txt_heigth)
                            lvc_var.set("")
                            ttk_labels.append(lfx3_c)
                        else:
                            sheet_name = "Faixa 3 Decrescente"
                            fx3_d = (f"  > Processando: {sheet_name}...")
                            lfx3_d = ttk.Label(page2, text=fx3_d, background="gray28")
                            lfx3_d.place(x=430, y=txt_heigth)
                            lvc_var.set("")
                            ttk_labels.append(lfx3_d)
                    elif file[-1] == "ADC" in file[-1]:
                        # FX3 ou ADC
                        if file[3] == "C":
                            sheet_name = "Faixa Adc Crescente"
                            fxa_c = (f"  > Processando: {sheet_name}...")
                            lfxa_c = ttk.Label(page2, text=fxa_c, background="gray28")
                            lfxa_c.place(x=430, y=txt_heigth)
                            lvc_var.set("")
                            ttk_labels.append(lfxa_c)
                        else:
                            sheet_name = "Faixa Adc Decrescente"
                            fxa_d = (f"  > Processando: {sheet_name}...")
                            lfxa_d = ttk.Label(page2, text=fxa_d, background="gray28")
                            lfxa_d.place(x=430, y=txt_heigth)
                            lvc_var.set("")
                            ttk_labels.append(lfxa_d)
                    else:
                        sheet_name = "Unknown"

                    txt_heigth += 20

                    root.update_idletasks()
                    page2.update_idletasks()

                    lvc_parameters = ld(lvc_path, intervalo, tipo)
                    lvc_df = lvc_parameters.run_command()
                    lvc_length = len(lvc_df)
                    lvc_progress.configure(maximum=lvc_length)
                    Fctn_Processamento.lvc_save(
                        dst_path, dst_path, lvc_df, sheet_name, 1, 9
                    )
                    lvc_progress["value"] = 0
                    lvc_var.set("")
                    root.update()
                    page2.update()
                    root.update_idletasks()
                    page2.update_idletasks()

            fim_do_processo = ttk.Label(page2, text=" ========================= FIM ========================", background="gray28")
            fim_do_processo.place(x=410, y=txt_heigth)
            ttk_labels.append(fim_do_processo)

            messagebox.showinfo("Aviso", "Planilha LVC DNIT exportada!")
            
            for label in ttk_labels:
                try:
                    label.destroy()
                except:
                    pass
                
            root.update()
            page2.update()
            root.update_idletasks()
            page2.update_idletasks()

    def RunIGG():
        path_igg = directory_entry.get()
        modelo_exportacao = model_igg.get()
        esp_igg = spin_igg.get()
        try:
            esp_igg = float(esp_igg)
        except:
            pass
        antt, artesp, dersp, pavesys, pavesys_faixa = False, False, False, False, False
        if modelo_exportacao == "ARTESP":
            artesp = True
        elif modelo_exportacao == "DER-SP":
            dersp = True
        elif modelo_exportacao == "ANTT":
            antt = True
        elif modelo_exportacao == "PAVESYS - Plataforma":
            pavesys = True
        elif modelo_exportacao == "PAVESYS - Faixa":
            pavesys_faixa = True

        if (
            (
                not antt
                and not artesp
                and not dersp
                and not pavesys
                and not pavesys_faixa
            )
            or (path_igg == "")
            or (esp_igg == "Espaçamento")
        ):
            messagebox.showerror(
                title="Erro de input",
                message="Todos os campos devem estar preenchidos!",
            )
        else:
            IGG(path_igg, antt, artesp, dersp, pavesys, pavesys_faixa, esp_igg)
            messagebox.showinfo("Aviso", "Planilha IGG exportada!")

    def RunLVDPavesys():
        path_lvd = directory_entry.get()
        if path_lvd == "":
            messagebox.showerror(
                title="Erro de input",
                message="Todos os campos devem estar preenchidos!",
            )
        else:
            BotaoLVDPavesys(path_lvd)
            messagebox.showinfo("Aviso", "Planilha LVD PAVESYS exportada!")


# ==============================================     PROCESS      ===============================================================#
def selectedItem():
    item_list = files_listbox.selection()
    if item_list:
        for item in item_list:
            file = files_listbox.item(item)["text"]
            if file not in process_listbox.get(0, tk.END):
                process_listbox.insert(tk.END, file)

        temp_list = sorted(process_listbox.get(0, tk.END))

        split_name = str(temp_list[0]).split("-")
        temp_name = str(
            split_name[0].replace("LVC 2m_", "")
            + "_D_"
            + split_name[1]
            + "_"
            + split_name[2]
        )
        temp_ini = float(split_name[3])
        try:
            temp_fim = float(
                str(temp_list[-1])
                .split("-")[-1]
                .replace(".xlsx", "")
                .replace(".xls", "")
            )
        except:
            temp_fim = float(
                str(temp_list[-1])
                .split("-")[-2]
                .replace(".xlsx", "")
                .replace(".xls", "")
            )

        ini_entry.delete(0, tk.END)
        fim_entry.delete(0, tk.END)
        name_spinbox.delete(0, tk.END)

        ini_entry.insert(0, temp_ini)
        fim_entry.insert(0, temp_fim)
        name_spinbox.insert(0, temp_name)


# ===============================================================================================================================#

root = tk.Tk()
root.resizable(False, False)

# CRIANDO A INTERFACE PRINCIPAL #
# ===============================================================================#
# DETERMINA O ESTILO, TAMANHO E TEMA
style = ttk.Style(root)
root.geometry("890x500")
root.tk.call("source", theme_path)
root.iconbitmap(icon_path)
root.title("Avaliações Interligadas - Pavesys")
style.theme_use("forest-dark")
# CRIA O FRAME PRINCIPAL
menu_frame = ttk.Frame(root, padding=10)
menu_frame.pack(side=tk.LEFT, fill=tk.Y)
# CRIA O FRAME DAS PAGINAS
pages_frame = ttk.Frame(root, padding=10)
pages_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
# CRIANDO PAGINAS
pages_notebook = ttk.Notebook(pages_frame)
# CRIAR PAGINA 1
page1 = ttk.Frame(pages_notebook, padding=10)
pages_notebook.add(page1, text="Processamento")
# CRIAR PAGINA 2
page2 = ttk.Frame(pages_notebook, padding=10)
pages_notebook.add(page2, text="      LVC DNIT      ")
# CRIAR PAGINA 3
page3 = ttk.Frame(pages_notebook, padding=10)
pages_notebook.add(page3, text="          IGG          ")
# CRIAR PAGINA 4
page4 = ttk.Frame(pages_notebook, padding=10)
pages_notebook.add(page4, text="LVD PAVESYS")
# CRIAR PAGINA 5
page5 = ttk.Frame(pages_notebook, padding=10)
pages_notebook.add(page5, text="     LVD DNIT     ")

pages_notebook.pack(fill=tk.BOTH, expand=True)

# ===============================================================================#
# CRIANDO PAGE1 >> PROCESSAMENTO
# Listagem de Planilhas
def toogle_mode():
    if switch_mode.instate(["selected"]):
        tipo_var.set(1)  # PADRAO
    else:
        tipo_var.set(0)  # DETALHADA


class new_window:
    def __init__(self) -> None:
        self.opennewwindow()
        pass

    def opennewwindow(self):
        global list_label

        self.newwindow = tk.Toplevel(root)
        self.newwindow.title("Ajuste de Lista")
        self.newwindow.geometry("320x500")
        self.newwindow.resizable(False, False)

        self.nw_tree = ttk.Treeview(
            self.newwindow, columns=("n°", "Início", "Fim"), height=17
        )
        self.nw_tree.heading("#0", text="n°")
        self.nw_tree.heading("#1", text="Início")
        self.nw_tree.heading("#2", text="Fim")
        self.nw_tree.column("#0", width=30, anchor="center")
        self.nw_tree.column("#1", width=120, anchor="center")
        self.nw_tree.column("#2", width=120, anchor="center")
        self.nw_tree.column("#3", width=1, anchor="center")
        self.nw_tree.grid(row=0, column=0, sticky="ew", columnspan=2, padx=5, pady=5)

        runlist_button = ttk.Button(
            self.newwindow,
            text="Gerar",
            command=lambda: Fctn_Processamento.run_pp_list(),
        )
        runlist_button.grid(row=2, column=0, sticky="ew", padx=5, pady=5)

        open_list = ttk.Button(
            self.newwindow,
            text="Abrir Lista",
            command=lambda: os.startfile(self.list_path),
        )
        open_list.grid(row=2, column=1, sticky="ew", padx=5, pady=5)

        list_label = ttk.Entry(self.newwindow)
        list_label.grid(row=1, column=0, sticky="ew", columnspan=2, padx=5, pady=5)

        self.values2nw()

    def list2values(self):
        main_path = os.getcwd()
        self.list_path = os.path.join(main_path, "cut_list.txt")

        df = pd.read_table(self.list_path, header=None)
        df.columns = ["Início", "Fim"]

        return df

    def values2nw(self):
        self.value_df = self.list2values()
        for index, row in self.value_df.iterrows():
            row_index = index + 1
            self.nw_tree.insert(
                "", "end", text=row_index, values=(row["Início"], row["Fim"])
            )

        name = name_spinbox.get()
        list_label.insert("0", name)

    def end(self):
        self.newwindow.destroy()

process_listbox = tk.Listbox(page1, width=50, height=10)
process_listbox.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=5)
files_listbox = ttk.Treeview(page1, columns=("Checkbox"), show="tree")
files_listbox.heading("#1", text="Item")
files_listbox.place(x=405, y=5)
pp_progress = ttk.Progressbar(page1, length=437)
pp_progress.place(x=406, y=300)
pp_var = tk.StringVar()
pp_var.set("")

percent_pp_label = ttk.Label(page1, textvariable=pp_var)
percent_pp_label.place(x=795, y=310)

tipo_var = tk.IntVar()
tipo_var.set(1)

switch_mode = ttk.Checkbutton(
    page1, text=" Padrão", style="Switch", command=toogle_mode, variable=tipo_var
)
switch_mode.place(x=170, y=325)
switch_label = ttk.Label(page1, text="Detalhada")
switch_label.place(x=99, y=326)

save_var = tk.StringVar()
save_var.set("")
save_label = ttk.Label(page1, textvariable=save_var)
save_label.place(x=406, y=310)
# Separador
separator = ttk.Separator(page1, orient="vertical")
separator.grid(row=0, column=2, rowspan=8, padx=10, sticky="ns")
# Entrada do Diretório
directory_entry = ttk.Entry(page1, width=40)
directory_entry.grid(row=6, column=3, sticky="ew", pady=5)
# Botão do Diretório
dir_button = ttk.Button(
    page1, text="Selecione Diretório", command=Fctn_Processamento.select_directory
)
dir_button.grid(row=6, column=5, padx=5, pady=5, sticky="ew")
# Botão Adicionar
add_button = ttk.Button(page1, text="<<", command=selectedItem)
add_button.place(x=405, y=250)
# Botão de Limpar
clean_button = ttk.Button(
    page1, text="Limpar", command=lambda: process_listbox.delete(0, tk.END)
)
clean_button.place(x=505, y=250)
# Botão de Concatenar Planilhas
concat_button = ttk.Button(
    page1, text="Concatenar", command=lambda: process_listbox.delete(0, tk.END)
)
concat_button.place(x=749, y=250)
# Botão de Corte por Lista
lista_button = ttk.Button(page1, text="Lista", command=new_window)
lista_button.place(x=649, y=250)
# Inicio
ini_label = ttk.Label(page1, text="km Início")
ini_label.grid(row=2, column=0, pady=5, sticky="ew")
ini_entry = ttk.Spinbox(page1, from_=0, to=1000, format="%10.3f", increment=0.01)
ini_entry.insert(0, 0.00)
ini_entry.grid(row=2, column=1, pady=5, sticky="ew")
# Fim
fim_label = ttk.Label(page1, text="km Fim")
fim_label.grid(row=3, column=0, pady=5, sticky="ew")
fim_entry = ttk.Spinbox(page1, from_=0, to=1000, format="%10.3f", increment=0.01)
fim_entry.insert(0, 100.00)
fim_entry.grid(row=3, column=1, pady=5, sticky="ew")
# nome
nome_label = ttk.Label(page1, text="Nome para Exportação")
nome_label.grid(row=4, column=0, pady=5, sticky="ew")
name_spinbox = ttk.Entry(page1)
name_spinbox.insert(0, "PP")
name_spinbox.grid(row=4, column=1, sticky="ew", pady=5)
# Botão de Gerar
run_button = ttk.Button(
    page1, text="Gerar", width=23, command=Fctn_Processamento.run_pre_processamento
)
run_button.grid(row=6, column=0, padx=0, sticky="w")
# Botão de Sair
exit_button = ttk.Button(
    page1, text="Sair", width=23, command=Fctn_Processamento.exit_application
)
exit_button.grid(row=6, column=0, columnspan=2, padx=5, sticky="e")
page1.rowconfigure(5, weight=1)
page1.columnconfigure(0, weight=0)
page1.columnconfigure(1, weight=0)
# ===============================================================================#
# CRIANDO PAGE2 >> LVC DNIT
# Listagem de Planilhas
files_listbox_lvc = tk.Listbox(page2, width=40, height=5)
files_listbox_lvc.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=5)
# Barra de Progresso
lvc_progress = ttk.Progressbar(page2, length=437)
lvc_progress.place(x=406, y=390)
lvc_var = tk.StringVar()
lvc_var.set("")
percent_lvc_label = ttk.Label(page2, textvariable=lvc_var)
percent_lvc_label.place(x=795, y=400)
# Separador
separator_lvc = ttk.Separator(page2, orient="vertical")
separator_lvc.grid(row=0, column=2, rowspan=8, padx=10, sticky="ns")
# Entrada do Diretório
directory_entry_lvc = ttk.Entry(page2, width=30)
directory_entry_lvc.grid(row=1, column=0, sticky="ew", pady=5)
# Botão do Diretório
dir_button_lvc = ttk.Button(
    page2, text="Selecione Diretório", command=Fctn_Processamento.select_directory
)
dir_button_lvc.grid(row=1, column=1, padx=5, pady=15, sticky="nsew")

# Processos
combo_list_lvc = ["PAVESYS"]
model_type_lvc = ttk.Combobox(page2, values=combo_list_lvc, width=50, state="readonly")
model_type_lvc.insert(0, "Modelos de Exportação")
model_type_lvc.grid(row=2, column=0, pady=15, columnspan=2, sticky="e")
# Texto
model_text_lvc = tk.Label(page2, text="MODELO:", font=("Arial", 9, "bold"), anchor="w")
model_text_lvc.place(x=1, y=153)
# SpinBox
sep_lvcdnit = ttk.Spinbox(page2, from_=0, to=1000, format="%10.0f", increment=100)
sep_lvcdnit.insert(0, "200")
sep_lvcdnit.bind("<FocusIn>", lambda e: sep_lvcdnit.delete("0", "end"))
sep_lvcdnit.bind("<Return>", Fctn_Processamento.on_enter_pressed)
sep_lvcdnit.grid(row=4, column=0, sticky="ew", pady=15)
# Texto
sep_text_lvc = tk.Label(page2, text="INTERVALO:", font=("Arial", 9, "bold"), anchor="w")
sep_text_lvc.place(x=1, y=215)
sep_metros_text_lvc = tk.Label(
    page2, text="[metros]", font=("Arial", 9, "bold"), anchor="w"
)
sep_metros_text_lvc.place(x=240, y=243)
# Botão de Gerar
run_button_lvc = ttk.Button(
    page2, text="Gerar", width=25, command=Fctn_Processamento.run_lvc_dnit
)
run_button_lvc.grid(row=6, column=0, pady=5, sticky="nsew", columnspan=2)

canvas = tk.Canvas(page2, width=430, height=360)
canvas.configure(bg="gray28")
canvas.place(x=407, y=1)

page2.rowconfigure(5, weight=1)
page2.columnconfigure(0, weight=0)
page2.columnconfigure(1, weight=0)
# ===============================================================================#

# CRIANDO PAGE3 >> IGG
# Listagem de Planilhas
files_listbox_igg = tk.Listbox(page3, width=40, height=5)
files_listbox_igg.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=5)
# Separador
separator_igg = ttk.Separator(page3, orient="vertical")
separator_igg.grid(row=0, column=2, rowspan=8, padx=10, sticky="ns")
# Entrada do Diretório
directory_entry_igg = ttk.Entry(page3, width=30)
directory_entry_igg.grid(row=1, column=0, sticky="ew", pady=5)
# Botão do Diretório
dir_button_igg = ttk.Button(
    page3, text="Selecione Diretório", command=Fctn_Processamento.select_directory
)
dir_button_igg.grid(row=1, column=1, padx=5, pady=15, sticky="nsew")
# Processos
list_igg = ["ARTESP", "DER-SP", "ANTT", "PAVESYS - Faixa", "PAVESYS - Plataforma"]
model_igg = ttk.Combobox(page3, values=list_igg, width=50, state="readonly")
model_igg.insert(0, "Modelos de Exportação")
model_igg.grid(row=2, column=0, pady=15, columnspan=2, sticky="e")
# Spinbox
spin_igg = ttk.Spinbox(page3, from_=0, to=1000, format="%10.3f", increment=0.01)
spin_igg.insert(0, "0.500")
spin_igg.bind("<FocusIn>", lambda e: spin_igg.delete("0", "end"))
spin_igg.grid(row=3, column=0, pady=15, sticky="ew")
# Botão de Gerar
run_button_igg = ttk.Button(
    page3, text="Gerar", width=25, command=Fctn_Processamento.RunIGG
)
run_button_igg.grid(row=6, column=0, pady=95, sticky="snew", columnspan=2)
# TEXTOS
model_text_igg = tk.Label(page3, text="MODELO:", font=("Arial", 9, "bold"), anchor="w")
model_text_igg.place(x=1, y=153)
sep_text_igg = tk.Label(page3, text="INTERVALO:", font=("Arial", 9, "bold"), anchor="w")
sep_text_igg.place(x=1, y=214)
sep_text_igg = tk.Label(page3, text="[kms]", font=("Arial", 9, "bold"), anchor="w")
sep_text_igg.place(x=240, y=246)
# ===============================================================================#

# CRIANDO PAGE4 >> LVD Pavesys
# Listagem de Planilhas
files_listbox_lvd = tk.Listbox(page4, width=40, height=5)
files_listbox_lvd.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=5)
# Separador
separator_lvd = ttk.Separator(page4, orient="vertical")
separator_lvd.grid(row=0, column=2, rowspan=8, padx=10, sticky="ns")
# Entrada do Diretório
directory_entry_lvd = ttk.Entry(page4, width=32)
directory_entry_lvd.grid(row=1, column=0, sticky="ew", pady=5)
# Botão do Diretório
dir_button_lvd = ttk.Button(
    page4, text="Selecione Diretório", command=Fctn_Processamento.select_directory
)
dir_button_lvd.grid(row=1, column=1, padx=5, pady=15, sticky="nsew")
# Botão de Gerar
run_button_lvd = ttk.Button(
    page4, text="Gerar", width=26, command=Fctn_Processamento.RunLVDPavesys
)
run_button_lvd.grid(row=3, column=0, pady=220, sticky="snew", columnspan=2)

# ===============================================================================#
# CRIANDO PAGE5 >> LVD DNIT
# Listagem de Planilhas
files_listbox_lvdd = tk.Listbox(page5, width=40, height=5)
files_listbox_lvdd.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=5)
# Separador
separator_lvdd = ttk.Separator(page5, orient="vertical")
separator_lvdd.grid(row=0, column=2, rowspan=8, padx=10, sticky="ns")
# Entrada do Diretório
directory_entry_lvdd = ttk.Entry(page5, width=32)
directory_entry_lvdd.grid(row=1, column=0, sticky="ew", pady=5)
# Botão do Diretório
dir_button_lvdd = ttk.Button(
    page5, text="Selecione Diretório", command=Fctn_Processamento.select_directory
)
dir_button_lvdd.grid(row=1, column=1, padx=5, pady=15, sticky="nsew")
# Botão de Gerar
run_button_lvdd = ttk.Button(
    page5, text="Gerar", width=26, command=Fctn_Processamento.RunLVDPavesys
)
run_button_lvdd.grid(row=3, column=0, pady=220, sticky="snew", columnspan=2)

root.mainloop()
